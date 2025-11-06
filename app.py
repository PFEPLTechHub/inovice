from functools import wraps
from flask_sqlalchemy import SQLAlchemy
from flask import Flask
from flask_cors import CORS
from flask import render_template, request, redirect, jsonify, session, url_for, send_file
import os
import io
import requests
from openpyxl import load_workbook
import random
import string
import pytz
from datetime import datetime, timedelta
from flask_mail import Message
from flask_mail import Mail
from flask_apscheduler import APScheduler
from static.python_functions import extras
from sqlalchemy import BigInteger
from sqlalchemy import text
import calendar
import pandas as pd
import re
import json
from functools import lru_cache
from flask import current_app

app = Flask(__name__)
app.secret_key = "eb9719b039ba725f3c1abe4224816938"

# Configure timeout settings to prevent connection resets
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Database Configuration (Replace with your actual DB URI)

#azeem creds
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:root@127.0.0.1:3306/invoice_system_testing'

# deployment server creds
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:root@127.0.0.1:3306/invoice_system'
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:azeem464693@127.0.0.1:3306/invoice_system_testing'

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Database connection pool settings to handle long-running queries
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,  # Verify connections before using them
    "pool_recycle": 3600,  # Recycle connections after 1 hour
    "pool_size": 10,  # Number of connections to maintain
    "max_overflow": 20,  # Maximum overflow connections
    "pool_timeout": 30,  # Timeout for getting connection from pool
}


app.config["MAIL_SERVER"] = "smtp.gmail.com"  # Change to your SMTP server
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "shaikhazeem4646@gmail.com"  # Change to your email
app.config["MAIL_PASSWORD"] = "mooicmexxptahzxo"  # Change to your password
app.config["MAIL_DEFAULT_SENDER"] = "shaikhazeem4646@gmail.com"  # Change to your email

# Initialize Mail
mail = Mail(app)

db = SQLAlchemy(app)

CORS(app)
# import secrets
# print(secrets.token_hex(16))  # Generates a random 32-character hex key


# --------------------------------------classes-----------------------------------------------------------------------------------


class contract_user_info(db.Model):
    __tablename__ = "contract_user_info"
    contractuserinfo_uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    employee_name = db.Column(db.String(255), nullable=False)
    employee_address = db.Column(db.String(255), nullable=False)
    employee_pincode = db.Column(BigInteger, nullable=False)
    pan_no = db.Column(db.String(20), nullable=False)
    account_number = db.Column(db.String(50), nullable=False)
    bank_name = db.Column(db.String(255), nullable=False)
    ifsc_code = db.Column(db.String(20), nullable=False)
    project_id = db.Column(db.Integer, nullable=False)
    monthly_salary = db.Column(db.Numeric(10, 2), nullable=False)
    food_allowance_per_day_amount = db.Column(db.Numeric(10, 2), nullable=False)
    phone_no = db.Column(db.String(15), nullable=False)
    mailid = db.Column(db.String(255), nullable=False)
    gender = db.Column(db.String(10), nullable=False)
    template_no = db.Column(db.String(50), nullable=False)
    email_send = db.Column(db.Boolean, nullable=False, default=False)
    joining_date = db.Column(db.Date, nullable=False)
    contractuserinfo_isactive = db.Column(db.Integer, nullable=False, default=1)
    manager_id = db.Column(
        db.Integer, db.ForeignKey("manager_info.managerinfo_uid"), nullable=True
    )
    # Optional per-employee company override (falls back to project's company_name)
    company_name = db.Column(db.String(255), nullable=True)
    company_address = db.Column(db.String(255), nullable=True)

    def to_dict(self):
        return {
            "contractuserinfo_uid": self.contractuserinfo_uid,
            "employee_name": self.employee_name,
            "employee_address": self.employee_address,
            "employee_pincode": self.employee_pincode,
            "pan_no": self.pan_no,
            "account_number": self.account_number,
            "bank_name": self.bank_name,
            "ifsc_code": self.ifsc_code,
            "project_id": self.project_id,
            "monthly_salary": float(self.monthly_salary),  # Convert Decimal to float
            "food_allowance_per_day_amount": float(self.food_allowance_per_day_amount),
            "phone_no": self.phone_no,
            "mailid": self.mailid,
            "gender": self.gender,
            "template_no": self.template_no,
            "email_send": self.email_send,
            "joining_date": self.joining_date.strftime("%Y-%m-%d"),  # Format Date
            "contractuserinfo_isactive": self.contractuserinfo_isactive,
            "manager_id": self.manager_id,
            "company_name": self.company_name,
            "company_address": self.company_address,
        }


class invoice_no_table(db.Model):
    __tablename__ = "invoice_no_table"

    invoicenotable_uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    last_invoice_number = db.Column(db.String(50), nullable=False)
    month_date = db.Column(db.String(100), nullable=True)
    payable_days = db.Column(db.Integer, nullable=True)
    food_amount = db.Column(db.Numeric(10, 2), nullable=True)
    arrears_month = db.Column(db.String(20), nullable=True)
    arrears_payable_days = db.Column(db.Integer, nullable=True)
    total_amount = db.Column(db.Numeric(10, 2), nullable=True)
    contract_employee_uid = db.Column(db.String(255), nullable=False)
    project_uid = db.Column(db.Integer, nullable=True)  # Add this new column
    invoicenotable_isactive = db.Column(db.Integer, nullable=False, default=1)
    manager_info_id = db.Column(db.Integer, nullable=True)
    last_date_generated = db.Column(db.TIMESTAMP, nullable=True, server_default=db.func.current_timestamp())
    yes_no = db.Column(db.String(3), nullable=True)
    city = db.Column(db.String(50))  # Stores invoice generation coordinates
    is_irregular = db.Column(db.Boolean, default=False)  # Flag for location mismatch


class history_log(db.Model):
    __tablename__ = "history_log"

    historylog_uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    invoice_no = db.Column(db.String(50), nullable=False)
    month_date = db.Column(db.String(50), nullable=False)
    contract_employee_uid = db.Column(db.Integer, nullable=False)
    historylog_isactive = db.Column(db.Integer, nullable=False, default=1)
    manager_info_id = db.Column(db.Integer, nullable=False)
    date = db.Column(
        db.DateTime, default=lambda: datetime.now(pytz.timezone("Asia/Kolkata"))
    )
    monthdate_empid = db.Column(db.String(255), nullable=True)
    entry_date_month = db.Column(db.String(50), nullable=True)
    location_irregular = db.Column(db.Boolean, default=False)  # Flag for location mismatch


class manager_info(db.Model):
    __tablename__ = "manager_info"

    managerinfo_uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    password = db.Column(db.String(255), nullable=False)
    isactive = db.Column(db.Integer, nullable=False, default=1)
    role = db.Column(db.Integer, nullable=False)
    manager_email = db.Column(db.String(255), unique=True, nullable=False)
    city = db.Column(db.String(50))


class projects_table(db.Model):
    __tablename__ = "projects_table"

    projectstable_uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(255), nullable=False)
    project_address = db.Column(db.String(255), nullable=False)
    company_name = db.Column(db.String(255), nullable=True, default="Pioneer Foundation Engineers Private Limited")
    isactive = db.Column(db.Integer, nullable=False, default=1)
    manager_id = db.Column(
        db.Integer, db.ForeignKey("manager_info.managerinfo_uid"), nullable=False
    )


class otp_validation(db.Model):
    __tablename__ = "otp_validation"

    uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    otp_no = db.Column(db.String(6), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    manager_id = db.Column(
        db.Integer, db.ForeignKey("manager_info.managerinfo_uid"), nullable=False
    )
    created_at = db.Column(db.TIMESTAMP, default=datetime.utcnow)

class InvoiceAutomationSettings(db.Model):
    __tablename__ = 'invoice_automation_settings'
    
    setting_id = db.Column(db.Integer, primary_key=True)
    is_enabled = db.Column(db.Integer, default=0)  # 0 = disabled, 1 = enabled
    default_month = db.Column(db.String(10))
    updated_by = db.Column(db.Integer)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
# --------------------------------------------------------------------------------------------------------------------------


# -----------------------------------------------routes----------------------------------------------------------------------


# Route protection: Ensure user is logged in
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "manager_data" not in session:
            return redirect(url_for("index"))  # Redirect to login if not authenticated
        return f(*args, **kwargs)

    return decorated_function


# Route protection: Ensure only admin users can access certain pages
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "manager_data" not in session or session["manager_data"].get("role") != 1:
            return redirect(url_for("index"))  # Redirect if not admin
        return f(*args, **kwargs)

    return decorated_function


@app.route("/")
def index():
    return render_template("loginForm.html")


@app.route("/invoice_main_page")
@login_required
def invoice_main_page():
    return render_template("invoice_main_page.html")


@app.route("/admin_homepage")
@admin_required
def admin_homepage():
    return render_template("admin_homepage.html")


@app.route("/history_homepage")
@login_required
def history_homepage():
    return render_template("history_homepage.html")


@app.route("/register_page")
@login_required
def register():
    return render_template("register_page.html")


@app.route("/project_table")
def get_projects():
    return render_template("project_table.html")


@app.route("/project_add")
def add_projects():
    return render_template("project_add.html")


@app.route("/user-information")
def user_informations():    
    return render_template("user_information.html")

@app.route('/report')
def report():
    """Render the report page with month pickers and table"""
    return render_template('report.html')

@app.route("/manager_page")
def manager_page():
    """Render the manager-employee page."""
    return render_template("manager_page.html")

@app.route("/bulk_upload")
def bulk_upload_page():
    return render_template("bulk_upload.html")

@app.route("/employee_payment_report")
def employee_payment_report_page():
    return render_template("employee_payment_report.html")

@app.route("/logout", methods=["POST"])
def logout():
    # Clear the session to log out the user
    session.clear()
    return redirect(url_for("index"))  # Redirect to login page






def get_user_roles():
    """Get the current user's role from session."""
    manager_data = session.get("manager_data")
    if not manager_data:
        return None
    return manager_data.get("role")


def get_current_user_id():
    """Get the current user's ID from session."""
    manager_data = session.get("manager_data")
    if not manager_data:
        return None
    return manager_data.get("manager_id")

@app.route('/get_automation_settings', methods=['GET'])
def get_automation_settings():
    """Get the current automation settings for invoice data."""
    try:
        # Only admins and managers should see this
        user_role = get_user_roles()
        if user_role not in [1, 0]: 
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Get latest settings
        settings = InvoiceAutomationSettings.query.order_by(
            InvoiceAutomationSettings.setting_id.desc()
        ).first()

        if not settings:
            return jsonify({
                'success': True,
                'settings': {
                    'is_enabled': 0,
                    'default_month': None
                }
            })

        return jsonify({
            'success': True,
            'settings': {
                'is_enabled': settings.is_enabled,
                'default_month': settings.default_month
            }
        })

    except Exception as e:
        print(f"Error getting automation settings: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/save_automation_settings', methods=['POST'])
def save_automation_settings():
    try:
        user_role = get_user_roles()
        if user_role != 1:
            return jsonify({'success': False, 'message': 'Unauthorized - Admin only'}), 403

        data = request.json
        is_enabled = data.get('is_enabled', 0)
        default_month = data.get('default_month')
        user_id = get_current_user_id()

        settings = InvoiceAutomationSettings.query.get(1)  # Fixed row
        if not settings:
            # Create the settings row if it doesn't exist
            settings = InvoiceAutomationSettings(
                setting_id=1,
                is_enabled=is_enabled,
                default_month=default_month,
                updated_by=user_id,
                updated_at=datetime.now()
            )
            db.session.add(settings)
        else:
            # Update the existing row
            settings.is_enabled = is_enabled
            settings.default_month = default_month
            settings.updated_by = user_id
            settings.updated_at = datetime.now()

        db.session.commit()
        return jsonify({'success': True, 'message': 'Settings updated successfully'})

    except Exception as e:
        db.session.rollback()
        print(f"Error updating automation settings: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/check_invoices_for_month', methods=['POST'])
def check_invoices_for_month():
    try:
        data = request.json
        month = data.get('month')
        if not month:
            return jsonify({'success': False, 'message': 'Month not provided'}), 400

        # Check if any invoice data exists for that month
        exists = db.session.query(
            db.session.query(invoice_no_table).filter(invoice_no_table.month_date == month).exists()
        ).scalar()

        if exists:
            return jsonify({'success': True, 'exists': True})
        else:
            return jsonify({'success': True, 'exists': False})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# -----------------------------------------------------------------------------------------------------------------------------------


@app.route("/get_managers_with_employees", methods=["GET"])
def get_managers_with_employees():
    # Get all managers with role=0
    managers = manager_info.query.filter_by(role=0, isactive=1).all()

    result = []

    for manager in managers:
        # Get all employees under this manager
        employees = contract_user_info.query.filter_by(
            manager_id=manager.managerinfo_uid, contractuserinfo_isactive=1
        ).all()

        # Create employee list
        employee_list = [
            {
                "employee_id": emp.contractuserinfo_uid,
                "employee_name": emp.employee_name,
                "monthly_salary": float(emp.monthly_salary),
                "joining_date": emp.joining_date.strftime("%Y-%m-%d"),
            }
            for emp in employees
        ]

        # Add manager with their employees to the result
        result.append(
            {
                "manager_id": manager.id,
                "manager_name": manager.name,
                "employees": employee_list,
                "employee_count": len(employee_list),
                'city': manager.city,
            }
        )

    return jsonify(result), 200

@app.route('/get_all_cities', methods=['GET'])
def get_all_cities():
    try:
        # Check if we have a cached result for the entire endpoint
        cached_result = getattr(get_all_cities, 'cached_result', None)
        if cached_result:
            return jsonify(cached_result)
            
        # Query all unique cities from the manager_info table
        managers = manager_info.query.filter_by(isactive=1).all()
        unique_cities = set()
        city_data = []
        
        for manager in managers:
            if manager.city and manager.city not in unique_cities:
                unique_cities.add(manager.city)
                # Get city name from cache or geocoding
                city_name = get_city_from_coordinates(manager.city)
                city_data.append({
                    "name": city_name,
                    "coordinates": manager.city
                })
        
        # Cache the result until the next server restart
        setattr(get_all_cities, 'cached_result', city_data)
        
        return jsonify(city_data)
    except Exception as e:
        print(f"Error fetching cities: {e}")
        return jsonify([]), 500

# Add a route to clear the cache if needed
@app.route('/admin/clear_city_cache', methods=['POST'])
def clear_city_cache():
    try:
        # Clear the in-memory caches
        get_city_from_coordinates.cache_clear()
        if hasattr(get_all_cities, 'cached_result'):
            delattr(get_all_cities, 'cached_result')
        
        # Clear the file cache
        global CITY_CACHE
        CITY_CACHE = {}
        with open(CITY_CACHE_FILE, 'w') as f:
            json.dump(CITY_CACHE, f)
            
        return jsonify({"message": "City cache cleared successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/update_manager_city', methods=['POST'])
def update_manager_city():
    try:
        data = request.json
        manager_id = data.get('manager_id')
        city = data.get('city')
        
        # Input validation
        if not manager_id or not city:
            return jsonify({"success": False, "message": "Missing required fields"}), 400
        
        # Update the manager's city
        manager = manager_info.query.filter_by(id=manager_id, isactive=1).first()
        
        if not manager:
            return jsonify({"success": False, "message": "Manager not found"}), 404
            
        manager.city = city
        db.session.commit()
        
        return jsonify({"success": True, "message": "City updated successfully"})
    
    except Exception as e:
        db.session.rollback()
        print(f"Error updating manager city: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# -----------------------------------------------------Login-------------------------------------------------------------------------
# Route to handle login
@app.route("/login", methods=["POST"])
def login():
    # Get the id and password from the request
    id = request.json.get("id")
    password = request.json.get("password")

    if not id or not password:
        return jsonify({"message": "Missing id or password"}), 400

    # Query the Manager table for the user by id
    manager = manager_info.query.filter_by(id=id, password=password).first()

    if manager:
        # Store manager info in session
        session["manager_data"] = {
            "manager_id": manager.managerinfo_uid,
            "role": manager.role,
        }

        manager_uid = manager.managerinfo_uid  # Extract UID

        # Query all projects managed by this manager
        projects = projects_table.query.filter_by(manager_id=manager_uid).all()

        # Extract project UIDs into a list
        project_list = [project.projectstable_uid for project in projects]

        # Store project list in session
        session["manager_data"]["project_list"] = project_list
        print("this is the project list", project_list)
        # Return success response based on role
        return jsonify({"success": 1 if manager.role == 1 else 0}), 200

    return jsonify({"message": "Invalid id or password"}), 401


# -----------------------------------------------invocie_data_workflow--------------------------------------------------------


@app.route("/api/manager-email", methods=["GET"])
def get_manager_email():
    # Check if user is logged in
    if "manager_data" not in session:
        return jsonify({"message": "Not logged in"}), 401

    # Get the manager ID from session
    manager_id = session.get("manager_data", {}).get("manager_id")

    if not manager_id:
        return jsonify({"message": "Manager ID not found"}), 404

    # Query the database to get the manager's email
    manager = manager_info.query.filter_by(managerinfo_uid=manager_id).first()

    if not manager or not hasattr(manager, "manager_email"):
        return jsonify({"message": "Manager email not found"}), 404

    # Return the email
    return jsonify({"email": manager.manager_email}), 200


# -----------------------------------------------------------------------------------------------------------------------------

@app.route("/user_information")
def user_information():
    # Check if user is logged in and is a manager
    if "manager_data" not in session:
        return redirect(url_for("login_page"))

    # Get manager info
    manager_uid = session["manager_data"]["manager_id"]
    manager = manager_info.query.filter_by(managerinfo_uid=manager_uid).first()

    if not manager or manager.role != 0:
        return redirect(url_for("/admin_homepage"))

    return render_template("user_information.html", manager=manager)


@app.route("/get_user_information")
def get_user_information():
    # Check if user is logged in and is a manager
    if "manager_data" not in session:
        return jsonify({"status": "error", "message": "Not logged in"})

    # Get manager info
    manager_uid = session["manager_data"]["manager_id"]
    manager = manager_info.query.filter_by(managerinfo_uid=manager_uid).first()

    if not manager or manager.role != 0:
        return jsonify({"status": "error", "message": "Unauthorized access"})

    # Get all employees managed by this manager
    employees = contract_user_info.query.filter_by(
        manager_id=manager_uid, contractuserinfo_isactive=1
    ).all()

    employee_list = []
    for emp in employees:
        employee_list.append(
            {
                "id": emp.contractuserinfo_uid,
                "employee_name": emp.employee_name,
                "employee_address": emp.employee_address,
                "project_id": emp.project_id,
                "phone_no": emp.phone_no,
                "mailid": emp.mailid,
                "gender": emp.gender,
                "template_no": emp.template_no,
                "joining_date": (
                    emp.joining_date.strftime("%Y-%m-%d") if emp.joining_date else ""
                ),
            }
        )

    return jsonify({"status": "success", "data": employee_list})


# -----------------------------------------------------
# Add these routes to your Flask application


@app.route("/generate_otp", methods=["POST"])
def generate_otp():
    # Check if user is logged in and is a manager
    if "manager_data" not in session:
        return jsonify({"status": "error", "message": "Not logged in"})

    # Parse request data
    data = request.get_json()
    employee_id = data.get("employee_id")

    if not employee_id:
        return jsonify({"status": "error", "message": "Employee ID is required"})

    # Get manager info
    manager_uid = session["manager_data"]["manager_id"]
    manager = manager_info.query.filter_by(managerinfo_uid=manager_uid).first()

    if not manager or manager.role != 0:  # Assuming role 0 is for managers
        return jsonify({"status": "error", "message": "Unauthorized access"})

    # Generate a random 6-digit OTP
    otp = "".join(random.choices(string.digits, k=6))

    # Deactivate any existing active OTPs for this manager
    db.session.query(otp_validation).filter(
        otp_validation.manager_id == manager_uid, otp_validation.is_active == 1
    ).update({otp_validation.is_active: 0})

    # Create a new OTP record
    new_otp = otp_validation(otp_no=otp, is_active=1, manager_id=manager_uid)

    db.session.add(new_otp)
    db.session.commit()

    # Send the OTP via email
    try:
        send_otp_email(manager.manager_email, otp)
        return jsonify({"status": "success", "message": "OTP sent successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to send OTP: {str(e)}"})


@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    # Check if user is logged in and is a manager
    if "manager_data" not in session:
        return jsonify({"status": "error", "message": "Not logged in"})

    # Parse request data
    data = request.get_json()
    employee_id = data.get("employee_id")
    otp = data.get("otp")

    if not employee_id or not otp:
        return jsonify({"status": "error", "message": "Missing required fields"})

    # Get manager info
    manager_uid = session["manager_data"]["manager_id"]
    manager = manager_info.query.filter_by(managerinfo_uid=manager_uid).first()

    if not manager or manager.role != 0:  # Assuming role 0 is for managers
        return jsonify({"status": "error", "message": "Unauthorized access"})

    # Check if the OTP is valid

    valid_otp = otp_validation.query.filter_by(
        manager_id=manager_uid, otp_no=otp, is_active=1
    ).first()

    if not valid_otp:
        return jsonify({"status": "error", "message": "Invalid or expired OTP"})

    # Mark the OTP as used
    valid_otp.is_active = 0
    db.session.commit()

    # Get the employee details
    employee = contract_user_info.query.filter_by(
        contractuserinfo_uid=employee_id,
        manager_id=manager_uid,
        contractuserinfo_isactive=1,
    ).first()

    if not employee:
        return jsonify({"status": "error", "message": "Employee not found"})

    # Return sensitive information
    sensitive_info = {
        "pan_no": employee.pan_no,
        "account_number": employee.account_number,
        "bank_name": employee.bank_name,
        "ifsc_code": employee.ifsc_code,
        "monthly_salary": float(employee.monthly_salary),
        "food_allowance_per_day_amount": float(employee.food_allowance_per_day_amount),
    }

    return jsonify({"status": "success", "data": sensitive_info})


# Helper function to send OTP via email
def send_otp_email(email, otp):
    try:
        msg = Message(
            "Your OTP for Sensitive Information Access",
            sender=app.config["MAIL_DEFAULT_SENDER"],
            recipients=[email],
        )

        msg.body = f"""
        Hello,
        
        Your One-Time Password (OTP) for accessing sensitive employee information is:
        
        {otp}
        
        This OTP is valid for 5 minutes.
        
        Please do not share this OTP with anyone.
        
        Regards,
        HR Management System
        """

        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Failed to send email: {str(e)}")
        raise e


# Scheduled task to expire OTPs (can be run with a scheduler like APScheduler)
def expire_old_otps():
    now = datetime.datetime.now()
    five_minutes_ago = now - datetime.timedelta(minutes=5)

    db.session.query(otp_validation).filter(
        otp_validation.is_active == 1, otp_validation.created_at < five_minutes_ago
    ).update({otp_validation.is_active: 0})

    db.session.commit()


scheduler = APScheduler()


def configure_scheduler(app):
    app.config["SCHEDULER_API_ENABLED"] = True
    scheduler.init_app(app)

    # Add job to expire old OTPs every minute
    scheduler.add_job(
        id="expire_otps", func=expire_old_otps, trigger="interval", minutes=1
    )

    scheduler.start()


# -------------------------------------------------------------------------------------------------------

@app.route("/get_initial_invoice_data", methods=["GET"])
def get_initial_invoice_data():
    # Retrieve manager_data from session
    manager_data = session.get("manager_data")
    print("invoicedata", manager_data)

    if not manager_data or "project_list" not in manager_data:
        return jsonify({"message": "Unauthorized access"}), 401

    manager_id = manager_data["manager_id"]
    project_list = manager_data["project_list"]
    user_role = manager_data.get("role", 0)  # Default to regular role if not specified

    # Query logic based on role
    if user_role == 1:  # Admin/Super user role
        data = contract_user_info.query.filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).order_by(contract_user_info.employee_name.asc()).all()
    else:  # Regular manager role
        data = contract_user_info.query.filter(
            contract_user_info.manager_id == manager_id,
            contract_user_info.contractuserinfo_isactive == 1,
        ).order_by(contract_user_info.employee_name.asc()).all()

    invoice_data = [
        {
            "contractuserinfo_uid": item.contractuserinfo_uid,
            "name": item.employee_name,
            "food_allowance_per_day_amount": item.food_allowance_per_day_amount,
            "monthly_salary": item.monthly_salary,
        }
        for item in data
    ]

    print(
        f"Fetched invoice data for {'all managers' if user_role == 1 else f'projects {project_list}'}:",
        len(invoice_data),
    )
    return jsonify(invoice_data)



# -------------------------------------------------------------------------------------------------------------------

@app.route("/search_invoices", methods=["POST"])
def search_invoices():
    data = request.json
    search_term = data.get("search_term", "")
    #selected_month = data.get("month", "")

    print("Received search term:", search_term)
    #print("Received month:", selected_month)

    #if not search_term or not selected_month:
        #return jsonify({"success": False, "message": "Missing search term or month"}), 400

    try:
        search_pattern = f"%{search_term}%"

        results = db.session.query(invoice_no_table.contract_employee_uid).join(
            contract_user_info,
            invoice_no_table.contract_employee_uid == contract_user_info.contractuserinfo_uid
        ).filter(
            #invoice_no_table.month_date == selected_month,
            invoice_no_table.invoicenotable_isactive == 1,
            db.or_(
                invoice_no_table.last_invoice_number.ilike(search_pattern),
                contract_user_info.employee_name.ilike(search_pattern),
                db.cast(invoice_no_table.payable_days, db.String).ilike(search_pattern),
                db.cast(invoice_no_table.food_amount, db.String).ilike(search_pattern),
                invoice_no_table.arrears_month.ilike(search_pattern),
                db.cast(invoice_no_table.total_amount, db.String).ilike(search_pattern),
            )
        ).all()

        uid_list = [row[0] for row in results]

        return jsonify({"success": True, "results": uid_list})

    except Exception as e:
        db.session.rollback()
        print(f"Error searching invoices: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500

# ---------------------------------------------------------------------------------------------------------------------
@app.route("/get_user_role", methods=["GET"])
def get_user_role():
    manager_data = session.get("manager_data", {})
    user_role = manager_data.get("role", 0)  # Default to 0 if no role is found
    print(f"DEBUG: get_user_role - manager_data={manager_data}, user_role={user_role}")
    return jsonify({"role": user_role})
# ------------------------------------------------------------------------------------------------------
@app.route("/get_employee_manager/<string:employee_id>", methods=["GET"])
def get_employee_manager(employee_id):
    try:
        emp = contract_user_info.query.get(employee_id)
        if not emp:
            return jsonify({"error": "Employee not found"}), 404
        return jsonify({"manager_id": emp.manager_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/update_employee_manager", methods=["POST"])
def update_employee_manager():
    try:
        data = request.get_json()
        employee_id = data.get("employee_id")
        manager_id = data.get("manager_id")

        if not employee_id or not manager_id:
            return jsonify({"success": False, "error": "Missing required fields"}), 400

        emp = contract_user_info.query.get(employee_id)
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404

        emp.manager_id = int(manager_id)
        db.session.commit()

        return jsonify({"success": True, "message": "Manager updated successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500



# ----------------------------------------------------------------------------------------------------------------------


@app.route("/get_managers1", methods=["GET"])
def get_managers1():
    """API endpoint to get all active managers for the dropdown (excluding role 1)."""
    try:
        # Query using the correct table name and column names
        # Filter out managers with role=1 and only get active managers
        managers = manager_info.query.filter(
            manager_info.isactive == 1, manager_info.role != 1
        ).all()

        managers_list = [
            {"id": manager.managerinfo_uid, "name": manager.name}
            for manager in managers
        ]
        return jsonify(managers_list), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/add_project", methods=["POST"])
def add_project():
    """API endpoint to add a new project to the database."""
    data = request.json
    project_name = data.get("project_name")
    project_address = data.get("project_address")
    manager_id = data.get("manager_id")
    isactive = data.get("isactive", 1)

    if not project_name or not project_address or not manager_id:
        return jsonify({"error": "Missing required fields"}), 400

    # Check if project with same name already exists
    existing_project = projects_table.query.filter_by(project_name=project_name).first()
    if existing_project:
        return (
            jsonify({"error": f"Project with name '{project_name}' already exists"}),
            400,
        )

    new_project = projects_table(
        project_name=project_name,
        project_address=project_address,
        manager_id=manager_id,
        isactive=isactive,
    )
    try:
        db.session.add(new_project)
        db.session.commit()
        return (
            jsonify({"message": f"Project '{project_name}' added successfully."}),
            201,
        )
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# -----------------------------------------------------------------------------------------

# Create a cache file path constant
CITY_CACHE_FILE = "city_cache.json"

# Load the cache from file on startup
try:
    with open(CITY_CACHE_FILE, 'r') as f:
        CITY_CACHE = json.load(f)
    print(f"Loaded {len(CITY_CACHE)} cities from cache")
except (FileNotFoundError, json.JSONDecodeError):
    CITY_CACHE = {}
    print("City cache not found or invalid, starting with empty cache")

@lru_cache(maxsize=1000)
def get_city_from_coordinates(coordinates):
    """
    Get city name from coordinates using cache first, then API if needed
    
    Args:
        coordinates (str): Comma-separated latitude and longitude (e.g., "19.0760,72.8777")
    Returns:
        str: City name or empty string if geocoding fails
    """
    # Clean up the coordinate string
    coordinates = coordinates.strip().replace('"', '').replace("'", "")
    
    # Check if coordinates are in cache
    if coordinates in CITY_CACHE:
        return CITY_CACHE[coordinates]
    
    try:
        # Verify coordinate format
        lat, lng = map(float, coordinates.split(','))
        
        # Using Nominatim API for reverse geocoding
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json"
        headers = {
            'User-Agent': 'InvoiceGenerationSystem/1.0'
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            city_name = ""
            # Extract city from address components
            if 'address' in data:
                for key in ['city', 'town', 'village', 'county', 'state_district']:
                    if key in data['address']:
                        city_name = data['address'][key]
                        break
            
            # Update the cache
            CITY_CACHE[coordinates] = city_name
            # Save the updated cache to file (could be optimized to not write on every update)
            try:
                with open(CITY_CACHE_FILE, 'w') as f:
                    json.dump(CITY_CACHE, f)
            except Exception as e:
                current_app.logger.error(f"Error saving city cache: {e}")
                
            return city_name
        else:
            current_app.logger.error(f"Geocoding API error: {response.status_code}")
            return ""
    except Exception as e:
        current_app.logger.error(f"Error in geocoding: {str(e)}")
        return ""

# ----------------------------------------------------------------------------------------------------------------------


@app.route("/submit_invoice", methods=["POST"])
def submit_invoice():
    manager_data = session.get("manager_data")
    if not request.is_json:
        return jsonify({"error": "Invalid content type, expected JSON"}), 415

    data = request.get_json()
    if not data:
        return jsonify({"error": "Empty request body"}), 400
        
    # Get current location coordinates from request
    current_coordinates = data.get("current_location", "")
    invoice_data = data.get("invoice_data", [])
    invoice_date = data.get("invoice_date", "")
    print("Received Data:", invoice_date)  # Debugging

    # Extract UIDs from invoice data
    uids = [row["uid"] for row in invoice_data]
    print("this is uid", uids)

    # Query database for matching rows
    matching_users = contract_user_info.query.filter(
        contract_user_info.contractuserinfo_uid.in_(uids)
    ).all()
    print("matching users", matching_users)
    # Ensure at least one user is found
    if not matching_users:
        return jsonify({"error": "No matching users found in database"}), 404

    # Extract manager_id from the first matching user
    manager_id = manager_data["manager_id"]
    print("dataaaaaaaaaaaaaa", manager_id)
    
    # Get manager's expected city coordinates
    manager_infos = manager_info.query.filter_by(managerinfo_uid=manager_id).first()
    if not manager_infos or not manager_infos.city:
        return jsonify({"error": "Manager location not configured"}), 400

        
    # Verify location if coordinates are provided
    location_match = True
    location_info = {}
    if current_coordinates:
        try:
        # Get city name from current coordinates through reverse geocoding
            current_city = get_city_from_coordinates(current_coordinates)
        # Get expected city from manager's coordinates
            expected_city = get_city_from_coordinates(manager_infos.city)

        
            location_match = (current_city.lower() == expected_city.lower())
            location_info = {
            "current_city": current_city,
            "expected_city": expected_city,
            "is_match": location_match
        }
            print("Current city:", current_city)
            print("Expected city:", expected_city)

        except Exception as e:
            print(f"Error in location verification: {str(e)}")
            # Continue with invoice generation even if location verification fails
        pass
  
    # Create a dictionary to map UID to the database row
    db_data_map = {str(user.contractuserinfo_uid): user for user in matching_users}
    print("this is the db data map", db_data_map)
    # Prepare final data to send to process_templates
    matched_data = []
    for row in invoice_data:
        uid = str(row["uid"])  # Ensure string type for consistency
        if uid in db_data_map:
            matched_data.append({
                "db_row": db_data_map[uid].to_dict(),  # Convert DB row to dictionary
                "table_row": row,  # Invoice row from request
            })

    # Pass manager_id and location info to process_templates
    output_folder = process_templates(matched_data, invoice_date, manager_id, current_coordinates, location_match)
    print("output folder", output_folder)
    # Zip the generated files
    zip_file_path = extras.zip_folder(output_folder, invoice_date)
    print("zip file path", zip_file_path)

    # Send the zip file to the frontend
    return send_file(zip_file_path, as_attachment=True)






def process_templates(matched_data, invoice_date, manager_id,current_coordinates="", location_match=True):
    print("Starting process_templates function")
    print("Manager ID:", manager_id)
    print("Invoice Date:", invoice_date)
    print("Matched Data Count:", len(matched_data))
    print("Current Coordinates:", current_coordinates)
    print("Location Match:", location_match)

    # Prepare output directory
    output_folder = "output"
    manager_folder = os.path.join(output_folder, str(manager_id))
    print("Manager Folder Path:", manager_folder)
    # Extract year and month from the invoice date
    formatted_date = invoice_date[:10]  # Format as 'YYYY-MM-DD'

    # Check if the folder for the month already exists and create a counter if needed
    invoice_folder = os.path.join(manager_folder, formatted_date)

    # If folder already exists, add a counter to the folder name
    counter = 1
    while os.path.exists(invoice_folder):
        invoice_folder = os.path.join(manager_folder, f"{formatted_date}_{counter}")
        counter += 1

    # Ensure the folder structure exists
    os.makedirs(invoice_folder, exist_ok=True)
    print("Final Invoice Folder Path:", invoice_folder)
    for index, data in enumerate(matched_data):
        db_row = data["db_row"]
        invoice_row = data["table_row"]
        print("Invoice Row Data:", invoice_row)

        # Convert SQLAlchemy row object to dictionary
        db_row_dict = dict(db_row)  # Fix: Directly use dictionary
        print("Database Row Data:", db_row_dict)
        print("this is the total salary", invoice_row["total_calculated_salary"])
        project_details = get_project_details(db_row_dict["project_id"])
        # Combine relevant data from db_row and invoice_row
        employee_data = {
            "table_uid": db_row_dict["contractuserinfo_uid"],
            "Employee_Name": db_row_dict["employee_name"],
            "Employee_Address": db_row_dict["employee_address"],
            "Employee_Pincode": db_row_dict["employee_pincode"],
            "PAN_No": db_row_dict["pan_no"],
            "Account_Number": db_row_dict["account_number"],
            "Bank_Name": db_row_dict["bank_name"],
            "IFSC_Code": db_row_dict["ifsc_code"],
            "Phone_Number": db_row_dict["phone_no"],
            "Email_Id": db_row_dict["mailid"],
            "Project_Name": project_details["name"],  # Get project name dynamically
            # Address shown on invoice: Pioneer -> project address; Others -> company address
            "Project_Address": (
                project_details["address"]
                if ((db_row_dict.get("company_name") or "Pioneer Foundation Engineers Private Limited") == "Pioneer Foundation Engineers Private Limited")
                else (db_row_dict.get("company_address") or "")
            ),
            # Use only employee-specific company name; if missing, use default
            "Company_Name": db_row_dict.get("company_name") or "Pioneer Foundation Engineers Private Limited",
            "Template_No": db_row_dict["template_no"].lower().strip(),
            "Monthly_Salary": db_row_dict["monthly_salary"],
            "Food_allowance_per_day_amount": db_row_dict[
                "food_allowance_per_day_amount"
            ],
            "project_id": db_row_dict["project_id"],
            # Invoice-specific data (from request)
            "Last_Invoice_No": invoice_row["invoice_no"],
            "Payable_Days": invoice_row["payable_days"],
            "Payable_Days_Amount": invoice_row["monthly_calculated_salary"],
            "Arrears_Month": invoice_row["arrears_month"],
            "Arrears_Days": invoice_row["arrears_payable_days"],
            "Arrears_Amount": invoice_row["arrears_amount"],
            "Food_Amount": invoice_row["food_amount_salary"],
            "Total_Salary": invoice_row["total_calculated_salary"],
            "Amount_in_words": invoice_row["amount_in_words"],
        }

        try:
            # Prepare template path
            template_no = employee_data["Template_No"]
            template_path = os.path.join("invoice_templates", f"{template_no}.xlsx")
            print("template path", template_path)

            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file '{template_path}' not found.")

            # Load the Excel template
            workbook = load_workbook(template_path)
            sheet = workbook.active  # Assuming the data is in the first sheet

            # Format invoice date
            employee_data["Last_Invoice_Date"] = invoice_date
            
            # Keep the original YYYY-MM format for database operations
            db_date_format = invoice_date[:7]  # Gets 'YYYY-MM' part
            
            # Convert to datetime object for display purposes only
            date_obj = datetime.strptime(db_date_format, "%Y-%m")
            # Format with month name for Excel display
            excel_formatted_date = date_obj.strftime("%B %Y")  # Will give "March 2025"
            
            arrears_month_str = employee_data["Arrears_Month"]
            
            # For arrears month, maintain both formats
            db_formatted_arrears_month = arrears_month_str  # Keep original for DB
            excel_formatted_arrears_month = None
            
            # Check if arrears month is provided and not empty
            if arrears_month_str:  # Ensure it's not None or an empty string
                try:
                    arrears_month_datetime = datetime.strptime(arrears_month_str, "%Y-%m")
                    excel_formatted_arrears_month = arrears_month_datetime.strftime("%B %Y")
                except ValueError:
                    excel_formatted_arrears_month = None  # Handle invalid date format gracefully
            
            formatted_invoice_date = datetime.strptime(invoice_date, "%Y-%m-%d").strftime("%d-%m-%Y")
            # Convert total salary to words
            print("this is the total salary", employee_data["Total_Salary"])
            # Ensure the value is converted to float or int before passing
            # Remove commas and convert to float
            total_salary = float(employee_data["Total_Salary"].replace(",", ""))
            total_amount_in_words = extras.format_amount_to_words(total_salary)
            employee_data["Amount_in_words"] = total_amount_in_words

            # Modify employee address for Excel formatting
            employee_address = f'{employee_data["Employee_Address"]} - {employee_data["Employee_Pincode"]}'
            project_address = get_project_address(employee_data["project_id"])
            if employee_data["Food_Amount"] == "":
                employee_data["Food_Amount"] = "0"
            print("Replacing placeholders in template")
            def format_decimal(value):
                try:
                    number = float(str(value).replace(",", "").strip())
                    if number.is_integer():
                        return str(int(number))  
                    else:
                        return str(number).rstrip('0').rstrip('.')  
                except:
                        return value  

            # Prepare placeholder map
            placeholder_map = {
    "{Enter Employee Name}": employee_data["Employee_Name"],
    "{Enter Employee Address}": employee_address,
    "{Enter PAN No.}": f'Pan No : {employee_data["PAN_No"]}',
    "{Enter Account Number}": f'Account Number : {employee_data["Account_Number"]}',
    "{Enter Bank Name}": f'Bank Name : {employee_data["Bank_Name"]}',
    "{Enter IFSC Code}": f'IFSC Code : {employee_data["IFSC_Code"]}',
    "{Enter Phone Number}": f'Mobile No :  {employee_data["Phone_Number"]}',
    "{Enter Email Id}": f'Email Id : {employee_data["Email_Id"]}',
    "{Enter Project Name}": f'Site : {employee_data["Project_Name"]}',
    "{Enter Project Address}": f'Address : {employee_data["Project_Address"]}',
    "{Enter Monthly Stipend}": employee_data["Monthly_Salary"],
    "{Enter Food allowance per day amount}": employee_data["Food_allowance_per_day_amount"],
    "{Enter Invoice No}": f'Invoice No : {employee_data["Last_Invoice_No"]}',
    "{Invoice Date}": f'Date : {formatted_invoice_date}     ',
    "{Month Year}": excel_formatted_date,  # Use display format for Excel
    "{Payable Days}": format_decimal(employee_data["Payable_Days"]),
    "{Payable Days Amount}": employee_data["Payable_Days_Amount"],
    "{Arrears Month}": excel_formatted_arrears_month,  # Use display format for Excel
    "{Arrears Days}": format_decimal(employee_data["Arrears_Days"]),
    "{Arrears Amount}": employee_data["Arrears_Amount"],
    "{Food Allowance Amount}": employee_data["Food_Amount"],
    "{Total Amount to be Credited}": employee_data["Total_Salary"],
    "{Total Amount to be Credited (in Words)}": f'Rupees : {employee_data["Amount_in_words"]}',
    # Company name replacement - this will replace hardcoded company names in templates
    "Pioneer Foundation Engineers Private Limited": employee_data["Company_Name"],
}
            for key, value in placeholder_map.items():
                print(f"Replacing {key} with {value}")
            # Replace placeholders in the template
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value in placeholder_map:
                        cell.value = placeholder_map[cell.value]

            # Define the base filename for the invoice
            base_filename = f"invoice_{employee_data['Employee_Name']}.xlsx"
            output_filename = os.path.join(invoice_folder, base_filename)

            # Check for uniqueness and add counter if needed
            counter = 1
            while os.path.exists(output_filename):
                output_filename = os.path.join(
                    invoice_folder,
                    f"invoice_{employee_data['Employee_Name']}_{counter}.xlsx",
                )
                counter += 1

            # Save the final invoice
            workbook.save(output_filename)
            workbook.close()
            print(f"Invoice saved: {output_filename}")

            extras.print_colored(
                f"\nTemplate created successfully for {index+1}. {employee_data['Employee_Name']}",
                "green",
            )

            emp_id = employee_data["table_uid"]
            invoice_no = employee_data["Last_Invoice_No"]
            print("this is the invoice no", invoice_no, emp_id, db_date_format)
            # Update DB with invoice data - use original format (YYYY-MM)
            update_invoice_no(emp_id, db_date_format, invoice_no, current_coordinates)
            add_history(emp_id, db_date_format, invoice_no, not location_match)

        except Exception as e:
            extras.log_text_to_file(f"Error while processing template: {str(e)}")
            extras.print_colored(f"Error while processing template: {str(e)}", "red")

    print("This is the invoice folder:", invoice_folder)
    return invoice_folder

# -------------------------------------------------------------------------------------------------------------------------------------


def get_project_address(uid):

    matching_project = projects_table.query.filter_by(
        projectstable_uid=uid, isactive=1
    ).first()
    print(
        "this is the name and the address",
        matching_project.project_name,
        matching_project.project_address,
    )


# --------------------------------------------------------------------------------------------------------------------------------------
def update_invoice_no(user_uid, date, invoice_no, current_coordinates=""):
    # Define IST timezone
    ist = pytz.timezone("Asia/Kolkata")

    # Get current time in IST
    current_time_ist = datetime.now(ist)

    query = invoice_no_table.query.filter_by(
        contract_employee_uid=user_uid, month_date=date
    ).first()
    # Fix: Use 'manager_id' key instead of 'managerinfo_uid'
    manager_info_id = session["manager_data"]["manager_id"]

    if not query:  # If no matching record is found, insert a new one
        new_entry = invoice_no_table(
            contract_employee_uid=user_uid,
            month_date=date,
            last_invoice_number=invoice_no,
            manager_info_id=manager_info_id,
            city=current_coordinates  # Store coordinates
        )
        db.session.add(new_entry)
        print("update", new_entry)
    else:  # If a record is found, update the existing one
        query.last_invoice_number = invoice_no  # Fix: use the correct column name
        query.last_date_generated = current_time_ist
        query.city = current_coordinates  # Update coordinates

    db.session.commit()  # Commit changes to the database



# --------------------------------------------------------------------------------------------------------------------------------------
def add_history(user_uid, date, invoice_no, wrong_location, location_irregular=False):
    # Define IST timezone
    ist = pytz.timezone("Asia/Kolkata")

    # Get current time in IST
    current_time_ist = datetime.now(ist)
    year_month = current_time_ist.strftime("%Y-%m")
    # Fix: Use 'manager_id' key instead of 'managerinfo_uid'
    manager_info_id = session["manager_data"]["manager_id"]
    monthdate_empid = f"{user_uid}|{date}"

    new_entry = history_log(
        contract_employee_uid=user_uid,
        month_date=date,
        invoice_no=invoice_no,
        manager_info_id=manager_info_id,
        date=current_time_ist,
        monthdate_empid=monthdate_empid,
        entry_date_month=year_month,
        location_irregular=location_irregular  # Flag irregular location
    )

    db.session.add(new_entry)
    db.session.commit()  # Commit changes to the database


# --------------------------------------------------------------------------------------------------------------------------------------


@app.route("/get_invoice_nos", methods=["POST"])
def get_invoice_nos():
    data = request.json
    month_date = data.get("month")  # Get selected date from frontend

    # Parse month_date to determine fiscal year
    try:
        year, month = month_date.split("-")
        year = int(year)
        month = int(month)

        # Determine fiscal year based on month
        if month >= 4:  # April and after
            fiscal_year = f"{year % 100}-{(year + 1) % 100}"
        else:  # Before April
            fiscal_year = f"{(year - 1) % 100}-{year % 100}"
    except:
        fiscal_year = "24-25"  # Default fiscal year if parsing fails

    invoice_data = {}

    # Fetch all user UIDs
    users = contract_user_info.query.order_by(contract_user_info.employee_name.asc()).all()

    for user in users:
        emp_uid = user.contractuserinfo_uid  # Get user UID

        # Check if there's an exact match in the history table for this user and month
        history_entry = invoice_no_table.query.filter_by(
            contract_employee_uid=emp_uid, month_date=month_date
        ).first()

        if history_entry:
            # If data exists for this month, fetch all fields
            invoice_data[emp_uid] = {
                "invoice_no": history_entry.last_invoice_number,
                "payable_days": float(history_entry.payable_days) if history_entry.payable_days else 0,
                "food_amount": float(history_entry.food_amount) if history_entry.food_amount else 0,
                "arrears_month": history_entry.arrears_month,
                "arrears_payable_days": history_entry.arrears_payable_days,
                "total_amount": float(history_entry.total_amount) if history_entry.total_amount else 0,
                "yes_no": history_entry.yes_no,  # Include the yes_no field
                "project_id": history_entry.project_uid  # Include project_uid
            }
        else:
            # Find the most recent invoice number for this fiscal year
            # First, format the date range for the current fiscal year
            if month >= 4:  # April to March of next year
                fiscal_start = f"{year}-04-01"
                fiscal_end = f"{year+1}-03-31"
            else:
                fiscal_start = f"{year-1}-04-01"
                fiscal_end = f"{year}-03-31"

            # Query for the latest invoice number in this fiscal year
            fiscal_year_invoices = (
                invoice_no_table.query.filter(
                    invoice_no_table.contract_employee_uid == emp_uid,
                    invoice_no_table.month_date >= fiscal_start,
                    invoice_no_table.month_date <= fiscal_end,
                )
                .order_by(invoice_no_table.last_date_generated.desc())
                .first()
            )

            # Get current project from contract_user_info table
            current_project_id = user.project_id

            if fiscal_year_invoices:
                # Found an invoice in the current fiscal year, increment its number
                parts = fiscal_year_invoices.last_invoice_number.split("/")
                if len(parts) == 2 and parts[1] == fiscal_year:
                    try:
                        invoice_num = int(parts[0])
                        # Increment the invoice number
                        next_num = invoice_num + 1
                        invoice_no = f"{next_num:02d}/{fiscal_year}"
                    except:
                        invoice_no = f"01/{fiscal_year}"
                else:
                    # Fiscal year doesn't match, start from 01
                    invoice_no = f"01/{fiscal_year}"
            else:
                # No invoices in this fiscal year, start from 01
                invoice_no = f"01/{fiscal_year}"

            # Only provide the invoice number for new entries
            invoice_data[emp_uid] = {
                "invoice_no": invoice_no,
                "payable_days": None,
                "food_amount": None,
                "arrears_month": None,
                "arrears_payable_days": None,
                "total_amount": None,
                "yes_no": "No",  # Default to "No" for new entries
                "project_id": current_project_id  # Include current project from contract_user_info
            }

    return jsonify(invoice_data)

# --------------------------------------------------------------------------------------------------------------------------------------
def get_last_invoice_no(emp_uid):
    """
    Fetch the last invoice number of a given employee UID.
    """
    last_invoice_entry = (
        invoice_no_table.query.filter_by(contract_employee_uid=emp_uid)
        .order_by(invoice_no_table.last_date_generated.desc())
        .first()
    )

    if last_invoice_entry:
        return last_invoice_entry.last_invoice_number
    else:
        return None  # No invoice found for this employee


# ------------------------------------------------------------------------------------------------------


# Route to get all active contract_user_info data
@app.route("/get_all_database_values", methods=["GET"])
def get_all_database_values():

    # Perform JOIN queries to get manager name and project name, ordered by employee name
    data = (
        db.session.query(
            contract_user_info.contractuserinfo_uid,
            contract_user_info.employee_name,
            contract_user_info.employee_address,
            contract_user_info.employee_pincode,
            contract_user_info.pan_no,
            contract_user_info.account_number,
            contract_user_info.bank_name,
            contract_user_info.ifsc_code,
            contract_user_info.project_id,
            contract_user_info.monthly_salary,
            contract_user_info.food_allowance_per_day_amount,
            contract_user_info.phone_no,
            contract_user_info.mailid,
            contract_user_info.gender,
            contract_user_info.template_no,
            contract_user_info.joining_date,
            manager_info.name.label("manager_name"),  # Fetch manager name
            projects_table.project_name.label("project_name"),  # Fetch project name
        )
        .outerjoin(
            manager_info, contract_user_info.manager_id == manager_info.managerinfo_uid
        )
        .outerjoin(
            projects_table,
            contract_user_info.project_id == projects_table.projectstable_uid,
        )
        .filter(contract_user_info.contractuserinfo_isactive == 1)
        .order_by(contract_user_info.employee_name.asc())  # Sort alphabetically
        .all()
    )

    result = []
    for item in data:
        # Get last invoice number for each employee
        last_invoice = get_last_invoice_no(item.contractuserinfo_uid)

        result.append(
            {
                "uid": item.contractuserinfo_uid,
                "name": item.employee_name,
                "last_invoice_no": last_invoice,
                "address": item.employee_address,
                "pincode": item.employee_pincode,
                "pan_no": item.pan_no,
                "account_number": item.account_number,
                "bank_name": item.bank_name,
                "ifsc_code": item.ifsc_code,
                "project_id": item.project_id,
                "project_name": (
                    item.project_name if item.project_name else "N/A"
                ),
                "monthly_salary": str(item.monthly_salary),
                "food_allowance": str(item.food_allowance_per_day_amount),
                "phone_no": item.phone_no,
                "mailid": item.mailid,
                "gender": item.gender,
                "template_no": item.template_no,
                "joining_date": item.joining_date.strftime("%Y-%m-%d"),
                "manager_name": item.manager_name,
            }
        )

    return jsonify(result)



# --------------------------------------------------------------------------------------------------------------------------------------

@app.route("/generate_excel", methods=["GET"])
def generate_excel():
    try:
        # Get employee data first
        data = get_all_database_values_for_excel()  # We'll create this function

        # Create directory structure
        today = datetime.now().strftime("%Y-%m-%d")
        base_path = os.path.join("static", "admin_homepage_excel", today)

        # Ensure directory exists
        os.makedirs(base_path, exist_ok=True)

        # Find next available counter
        counter = 1
        while True:
            file_name = f"{today}_{counter}.xlsx"
            file_path = os.path.join(base_path, file_name)
            if not os.path.exists(file_path):
                break
            counter += 1

        # Convert data to DataFrame
        df = pd.DataFrame(data)

        # Write to Excel
        df.to_excel(file_path, index=False)

        # Return success with file path for download
        return jsonify(
            {
                "success": True,
                "message": f"Excel generated successfully: {file_name}",
                "file_path": file_path,
            }
        )

    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Error: {str(e)}"})


# Helper function to get formatted data for Excel
def get_all_database_values_for_excel():
    # Get the data from the database with joins
    data = (
        db.session.query(
            contract_user_info.contractuserinfo_uid,
            contract_user_info.employee_name,
            contract_user_info.employee_address,
            contract_user_info.employee_pincode,
            contract_user_info.pan_no,
            contract_user_info.account_number,
            contract_user_info.bank_name,
            contract_user_info.ifsc_code,
            contract_user_info.project_id,
            contract_user_info.monthly_salary,
            contract_user_info.food_allowance_per_day_amount,
            contract_user_info.phone_no,
            contract_user_info.mailid,
            contract_user_info.gender,
            contract_user_info.template_no,
            contract_user_info.joining_date,
            manager_info.name.label("manager_name"),
            projects_table.project_name.label("project_name"),
        )
        .outerjoin(
            manager_info, contract_user_info.manager_id == manager_info.managerinfo_uid
        )
        .outerjoin(
            projects_table,
            contract_user_info.project_id == projects_table.projectstable_uid,
        )
        .filter(contract_user_info.contractuserinfo_isactive == 1).order_by(
    contract_user_info.employee_name.asc()  # Sort alphabetically
).all()
    )

    result = []
    for index, item in enumerate(data):
        # Get last invoice number for each employee
        last_invoice = get_last_invoice_no(item.contractuserinfo_uid)

        result.append(
            {
                "S.No": index + 1,
                "Name": item.employee_name,
                "Last Invoice No": last_invoice if last_invoice else "N/A",
                "Address": item.employee_address,
                "Pin Code": item.employee_pincode,
                "PAN No": item.pan_no,
                "Account Number": item.account_number,
                "Bank Name": item.bank_name,
                "IFSC Code": item.ifsc_code,
                "Project": item.project_name if item.project_name else "N/A",
                "Monthly Salary": str(item.monthly_salary),
                "Food Allowance": str(item.food_allowance_per_day_amount),
                "Phone No": item.phone_no,
                "Email": item.mailid,
                "Gender": item.gender,
                "Template No": item.template_no,
                "Joining Date": (
                    item.joining_date.strftime("%Y-%m-%d") if item.joining_date else ""
                ),
                "Manager Name": item.manager_name if item.manager_name else "N/A",
            }
        )

    return result


# Add route to download the generated Excel file
@app.route("/", methods=["GET"])
def download_excel(filename):
    directory = os.path.join("static", "admin_homepage_excel")
    return send_file(os.path.join(directory, filename), as_attachment=True)
#---------------------------------------------------------------------------------------------------------------------------------------


# ---------------------------------------------------------------------------------------------------------------------------------------
# Define the route to handle the delete request
@app.route("/delete_values", methods=["POST"])
def delete_values():
    # Get the JSON data sent from the front end
    request_data = request.get_json()

    # Extract the uid from the request data
    uid_to_delete = request_data.get("uid")

    # Print the UID to the console
    print("Received UID to delete:", uid_to_delete)

    record = contract_user_info.query.filter_by(
        contractuserinfo_uid=uid_to_delete
    ).first()
    print(f"value of record is: {record}")

    if record:
        print("ayyyeyyeeee")
        # Update the 'value_softDel' column to 0
        record.contractuserinfo_isactive = 0

        # db.session.delete(record)
        # Commit the changes to the database
        db.session.commit()

        # # Optionally, refresh the record to ensure the update is applied
        # db.session.refresh(record)

        # Respond back to the front end
        return jsonify({"success": "Record updated successfully", "uid": uid_to_delete})
    else:
        return jsonify({"message": "UID not found", "uid": uid_to_delete}), 404


# --------------------------------------------------------------------------------------------------------------------------------------
@app.route("/api/invoices", methods=["GET"])
def get_all_invoices():
    try:
        # Query all active invoices from the database
        invoices = invoice_no_table.query.filter_by(invoicenotable_isactive=1).all()

        # Print the fetched invoices to the console
        print("Fetched invoices:", invoices)

        # Convert the SQLAlchemy objects to dicts
        invoice_list = []
        for invoice in invoices:
            invoice_data = {
                "id": invoice.invoicenotable_uid,
                "number": invoice.invoicenotable_number,
                "date": invoice.invoicenotable_date,
                "isactive": invoice.invoicenotable_isactive,
                # Add any other fields as necessary
            }
            invoice_list.append(invoice_data)

        # Return the data as JSON
        return jsonify({"status": "success", "data": invoice_list}), 200

    except Exception as e:
        # Print the error to the console
        print("Error fetching invoices:", e)

        # Return error response
        return jsonify({"status": "error", "message": "Failed to fetch invoices"}), 500


@app.route("/update_values", methods=["POST"])
def update_values():
    data = request.json
    uid = data.get("uid")
    project_id = data.get("project_id")  # Extract project_id from request

    print("Received UID:", uid)
    print("Received Data:", data)

    if not uid:
        return jsonify({"success": False, "error": "UID is required"}), 400

    # Fetch user from the database
    user = contract_user_info.query.filter_by(contractuserinfo_uid=uid).first()
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 404

    update_data = {
        key: value for key, value in data.items() if key not in ["uid", "project_id"]
    }

    if not update_data and not project_id:
        return jsonify({"success": False, "error": "No data provided to update"}), 400

    try:
        # Update other user attributes
        for key, value in update_data.items():
            if hasattr(user, key):
                setattr(user, key, value)

        # Explicitly update project_id if provided
        if project_id:
            user.project_id = project_id
            print(f"Updated project_id to {project_id}")

        db.session.commit()
        return jsonify({"success": True, "message": "Data updated successfully"})

    except Exception as e:
        db.session.rollback()
        print(f"Error updating user: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


# --------------------------------------------------------------------------------------------------------------------------------------

@app.route("/history_homepage_initial_data")
def history_homepage_initial_data():
    print("we are hereeee")
    
    # Get manager data from session for filtering
    manager_data = session.get("manager_data")
    if not manager_data:
        return jsonify({"message": "Unauthorized access"}), 401
    
    manager_id = manager_data["manager_id"]
    user_role = manager_data.get("role", 0)
    
    print(f"DEBUG: history_homepage_initial_data - manager_id={manager_id}, user_role={user_role}")

    # Get the current and previous month in YYYY-MM format
    current_month = datetime.now().strftime("%Y-%m")
    previous_month = (datetime.now().replace(day=1) - timedelta(days=1)).strftime(
        "%Y-%m"
    )

    # Fetch data for only the current and previous month
    query = db.session.query(history_log)
    
    # Filter by manager if user is not admin
    # Convert to int to handle string comparison issues
    user_role = int(user_role) if user_role is not None else 0
    if user_role != 1:  # Not admin, filter by manager
        # Join with contract_user_info to filter by manager_id
        query = query.join(
            contract_user_info,
            history_log.contract_employee_uid == contract_user_info.contractuserinfo_uid
        ).filter(
            contract_user_info.manager_id == manager_id,
            history_log.entry_date_month.in_([current_month, previous_month])
        )
    else:  # Admin can see all data
        query = query.filter(history_log.entry_date_month.in_([current_month, previous_month]))
    
    query = query.order_by(history_log.month_date.asc()).all()

    data = [
        {
            "historylog_uid": item.historylog_uid,
            "invoice_no": item.invoice_no,
            "month_date": item.month_date,
            "contract_employee_uid": item.contract_employee_uid,
            "historylog_isactive": item.historylog_isactive,
            "manager_info_id": item.manager_info_id,
            "date": item.date,
            "monthdate_empid": item.monthdate_empid,
            "location_irregular": item.location_irregular,
        }
        for item in query
    ]

    df = pd.DataFrame(data)

    # Remove duplicates based on 'monthdate_empid', keeping the last occurrence
    filtered_df = df[~df.duplicated("monthdate_empid", keep="last")]

    history_list = []
    for _, record in filtered_df.iterrows():
        employee = contract_user_info.query.filter_by(
            contractuserinfo_uid=record["contract_employee_uid"]
        ).first()
        employee_name = employee.employee_name if employee else "Unknown"

        manager = manager_info.query.filter_by(
            managerinfo_uid=record["manager_info_id"]
        ).first()
        manager_name = manager.name if manager else "Unknown"

        history_list.append(
            {
                "historylog_uid": record["historylog_uid"],
                "invoice_no": record["invoice_no"],
                "month_date": record["month_date"],
                "contract_employee_uid": record["contract_employee_uid"],
                "contract_employee_name": employee_name,
                "manager_info_id": record["manager_info_id"],
                "manager_name": manager_name,
                "date": (
                    record["date"].strftime("%d-%m-%Y %H:%M:%S")
                    if isinstance(record["date"], datetime)
                    else "Unknown"
                ),
                "location_irregular": record["location_irregular"],
            }
        )

    # Return both the table data & all available months and invoice numbers
    return jsonify(
        {
            "history_list": history_list,  # All data sorted by month_date
        }
    )


# --------------------------------------------------------------------------------------------------------------------------------------


@app.route("/get_history_month_filtered_data", methods=["POST"])
def get_history_month_filtered_data():
    data = request.json
    selected_month = data.get("month", "")
    show_duplicates = data.get("selected_button", "") == "true"
    
    # Get manager data from session for filtering
    manager_data = session.get("manager_data")
    if not manager_data:
        return jsonify({"message": "Unauthorized access"}), 401
    
    manager_id = manager_data["manager_id"]
    user_role = manager_data.get("role", 0)

    # Query database for entries matching the selected month
    query = db.session.query(history_log)
    
    # Filter by manager if user is not admin
    # Convert to int to handle string comparison issues
    user_role = int(user_role) if user_role is not None else 0
    if user_role != 1:  # Not admin, filter by manager
        # Join with contract_user_info to filter by manager_id
        query = query.join(
            contract_user_info,
            history_log.contract_employee_uid == contract_user_info.contractuserinfo_uid
        ).filter(contract_user_info.manager_id == manager_id)
    
    if selected_month:
        query = query.filter(history_log.month_date == selected_month)

    # Convert query results to a list of dictionaries
    raw_data = [
        {
            "historylog_uid": item.historylog_uid,
            "invoice_no": item.invoice_no,
            "month_date": item.month_date,
            "contract_employee_uid": item.contract_employee_uid,
            "historylog_isactive": item.historylog_isactive,
            "manager_info_id": item.manager_info_id,
            "date": item.date,
            "monthdate_empid": item.monthdate_empid,
            "location_irregular": item.location_irregular,
        }
        for item in query.all()
    ]

    # If "Show Duplicates" is enabled, return all rows
    if show_duplicates:
        filtered_data = raw_data
    else:
        # Convert to Pandas DataFrame and remove duplicates
        df = pd.DataFrame(raw_data)
        filtered_df = df[~df.duplicated("monthdate_empid", keep="last")]
        filtered_data = filtered_df.to_dict(orient="records")

    # Fetch employee and manager names for each record
    history_list = []
    for record in filtered_data:
        employee = contract_user_info.query.filter_by(
            contractuserinfo_uid=record["contract_employee_uid"]
        ).first()
        employee_name = employee.employee_name if employee else "Unknown"

        manager = manager_info.query.filter_by(
            managerinfo_uid=record["manager_info_id"]
        ).first()
        manager_name = manager.name if manager else "Unknown"

        history_list.append(
            {
                "historylog_uid": record["historylog_uid"],
                "invoice_no": record["invoice_no"],
                "month_date": record["month_date"],
                "contract_employee_uid": record["contract_employee_uid"],
                "contract_employee_name": employee_name,
                "manager_info_id": record["manager_info_id"],
                "manager_name": manager_name,
                "date": (
                    record["date"].strftime("%d-%m-%Y %H:%M:%S")
                    if isinstance(record["date"], datetime)
                    else "Unknown"
                ),
                "location_irregular": record["location_irregular"],
            }
        )

    return jsonify(history_list)


# ------------------------------------------------------------------------------------------------------
@app.route('/history_employees', methods=['GET'])
def history_employees():
    try:
        # Get manager data from session for filtering
        manager_data = session.get("manager_data")
        if not manager_data:
            return jsonify({"message": "Unauthorized access"}), 401
        
        manager_id = manager_data["manager_id"]
        user_role = manager_data.get("role", 0)
        
        print(f"DEBUG: manager_id={manager_id}, user_role={user_role}")
        
        # Filter employees by manager if user is not admin
        # Convert to int to handle string comparison issues
        user_role = int(user_role) if user_role is not None else 0
        if user_role != 1:  # Not admin, filter by manager
            employees = contract_user_info.query.filter_by(
                contractuserinfo_isactive=1,
                manager_id=manager_id
            ).order_by(contract_user_info.employee_name.asc()).all()
            print(f"DEBUG: Found {len(employees)} employees for manager {manager_id}")
        else:  # Admin can see all employees
            employees = contract_user_info.query.filter_by(contractuserinfo_isactive=1).order_by(contract_user_info.employee_name.asc()).all()
            print(f"DEBUG: Admin access - Found {len(employees)} total employees")
        
        return jsonify([
            {"id": e.contractuserinfo_uid, "name": e.employee_name}
            for e in employees
        ])
    except Exception as e:
        print(f"DEBUG: Error in history_employees: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/history_employee_fy', methods=['POST'])
def history_employee_fy():
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        if not employee_id:
            return jsonify({"history_list": []})
        
        # Get manager data from session for filtering
        manager_data = session.get("manager_data")
        if not manager_data:
            return jsonify({"message": "Unauthorized access"}), 401
        
        manager_id = manager_data["manager_id"]
        user_role = manager_data.get("role", 0)

        # Determine current financial year start (April 1 of current FY)
        today = datetime.now().date()
        fy_year = today.year if today.month >= 4 else today.year - 1
        fy_start_month = f"{fy_year}-04"

        # Fetch all history rows for this employee from FY start to now
        q = db.session.query(history_log).filter(
            history_log.contract_employee_uid == str(employee_id),
            history_log.month_date >= fy_start_month
        )
        
        # Additional security: ensure manager can only see their own employees
        # Convert to int to handle string comparison issues
        user_role = int(user_role) if user_role is not None else 0
        if user_role != 1:  # Not admin, verify employee belongs to manager
            q = q.join(
                contract_user_info,
                history_log.contract_employee_uid == contract_user_info.contractuserinfo_uid
            ).filter(contract_user_info.manager_id == manager_id)
        
        q = q.order_by(history_log.month_date.asc(), history_log.invoice_no.asc())

        rows = q.all()

        history_list = []
        for item in rows:
            # Join to names
            employee = contract_user_info.query.filter_by(contractuserinfo_uid=item.contract_employee_uid).first()
            employee_name = employee.employee_name if employee else "Unknown"
            manager = manager_info.query.filter_by(managerinfo_uid=item.manager_info_id).first()
            manager_name = manager.name if manager else "Unknown"

            history_list.append({
                "historylog_uid": item.historylog_uid,
                "invoice_no": item.invoice_no,
                "month_date": item.month_date,
                "contract_employee_uid": item.contract_employee_uid,
                "contract_employee_name": employee_name,
                "manager_info_id": item.manager_info_id,
                "manager_name": manager_name,
                "date": (item.date.strftime("%d-%m-%Y %H:%M:%S") if isinstance(item.date, datetime) else "Unknown"),
                "location_irregular": item.location_irregular,
            })

        return jsonify({"history_list": history_list})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --------------------------------------------------------------------------------------------------------------------------------------
@app.route("/register", methods=["POST"])
def register_user():
    data = request.json

    # Create a new ContractUserInfo entry with project_id and manager_id set to 1
    new_user = contract_user_info(
        employee_name=data["employee_name"],
        employee_address=data["employee_address"],
        employee_pincode=data["employee_pincode"],
        pan_no=data["pan_no"],
        account_number=data["account_number"],
        bank_name=data["bank_name"],
        ifsc_code=data["ifsc_code"],
        monthly_salary=data["monthly_salary"],
        food_allowance_per_day_amount=data.get("food_allowance_per_day_amount") or 0,
        phone_no=data["phone_no"],
        mailid=data["mailid"],
        gender=data["gender"],
        template_no=data["template_no"],
        joining_date=data["joining_date"],
        project_id=data["project_name"], 
        manager_id=data["manager_name"],
        company_name=data.get("company_name") or None,
        company_address=data.get("company_address") or None,
    )

    db.session.add(new_user)
    db.session.commit()

    return (
        jsonify(
            {
                "message": "User registered successfully",
                "user_id": new_user.contractuserinfo_uid,
            }
        ),
        201,
    )


# --------------------------------------------------------------------------------------------------------------------------------------


# Route to get all active get_register_project data
@app.route("/get_register_project_names", methods=["GET"])
def get_project_names():
    projects = projects_table.query.filter_by(isactive=1).all()
    project_dict = {
        project.projectstable_uid: project.project_name for project in projects
    }
    return jsonify(project_dict)


@app.route("/get_register_project_names_admin", methods=["GET"])
def get_project_names_admin():
    projects = (
        db.session.query(
            projects_table.projectstable_uid,
            projects_table.project_name,
            manager_info.name.label("manager_name"),
        )
        .join(manager_info, projects_table.manager_id == manager_info.managerinfo_uid)
        .filter(projects_table.isactive == 1)
        .all()
    )

    project_list = [
        {"uid": p.projectstable_uid, "name": f"{p.project_name} ({p.manager_name})"}
        for p in projects
    ]

    return jsonify(project_list)


@app.route("/get_register_project_names", methods=["GET"])
def get_project_details(project_id):
    try:
        # Query the projects table for the specific project
        project = projects_table.query.filter_by(
            projectstable_uid=project_id, isactive=1
        ).first()

        if project:
            return {
                "name": project.project_name,
                "address": project.project_address,
            }
        return {"name": "Project Not Found", "address": "Address Not Found"}
    except Exception as e:
        print(f"Error fetching project details: {str(e)}")
        return {"name": "Project Not Found", "address": "Address Not Found"}


# --------------------------------------------------------------------------------------------------------------------------------------
@app.route("/get_projects", methods=["GET"])
def get_project():
    print("Fetching projects from the database...")  # Debugging statement

    projects = (
        db.session.query(
            projects_table.projectstable_uid,
            projects_table.project_name,
            projects_table.project_address,
            projects_table.manager_id,  # This stores managerinfo_uid
            manager_info.name.label("manager_name"),
        )
        .join(manager_info, projects_table.manager_id == manager_info.managerinfo_uid)
        .all()
    )

    print(f"Number of projects found: {len(projects)}")  # Debugging statement

    project_list = [
        {
            "project_id": project.projectstable_uid,  # Add project_id for updates
            "project_name": project.project_name,
            "project_address": project.project_address,
            "manager_id": project.manager_id,  # Include manager_id
            "manager_name": project.manager_name,
        }
        for project in projects
    ]

    print("Project list:", project_list)  # Debugging statement

    return jsonify(project_list), 200


@app.route("/get_register_project_names_display", methods=["GET"])
def get_register_project_names_display():
    """Return projects for dropdown as id/name with manager name appended."""
    try:
        projects = (
            db.session.query(
                projects_table.projectstable_uid,
                projects_table.project_name,
                manager_info.name.label("manager_name"),
            )
            .join(manager_info, projects_table.manager_id == manager_info.managerinfo_uid)
            .filter(projects_table.isactive == 1)
            .all()
        )

        result = [
            {
                "id": p.projectstable_uid,
                "name": f"{p.project_name} ({p.manager_name})",
            }
            for p in projects
        ]
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/get_employee_project/<int:employee_id>", methods=["GET"])
def get_employee_project(employee_id):
    try:
        employee = contract_user_info.query.filter_by(contractuserinfo_uid=employee_id).first()
        if employee:
            return jsonify({
                "project_id": employee.project_id,
            }), 200
        return jsonify({"error": "Employee not found"}), 404
    except Exception as e:
        print(f"Error fetching employee project: {str(e)}")
        return jsonify({"error": str(e)}), 500



# --------------------------------------------------------------------------------------------------------------------------------------


@app.route("/get_managers", methods=["GET"])
def get_managers():
    managers = manager_info.query.filter_by(
        role=0
    ).all()  # Fetch only managers with role = 0
    manager_list = [
        {"manager_id": manager.managerinfo_uid, "manager_name": manager.name}
        for manager in managers
    ]
    return jsonify(manager_list), 200


# --------------------------------------------------------------------------------------------------------------------------------------


@app.route("/update_project_manager", methods=["POST"])
def update_project_manager():
    data = request.json
    project_id = data.get("project_id")
    new_manager_id = data.get("manager_id")

    if not project_id or not new_manager_id:
        return jsonify({"error": "Missing project_id or manager_id"}), 400

    project = projects_table.query.get(project_id)
    if not project:
        return jsonify({"error": "Project not found"}), 404

    # Update the manager_id for the selected project
    project.manager_id = new_manager_id
    db.session.commit()

    return jsonify({"message": "Project manager updated successfully"}), 200


# --------------------------------------------------------------------------------------------------------------------------------------


# @app.route('/generate_invoice_no', methods=['POST'])
def generate_invoice_no(current_invoice):
    try:
        parts = current_invoice.split("/")

        # Check if the current_invoice format is valid
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].count("-") == 1:
            extras.print_colored(f"Invalid invoice format: {current_invoice}", "red")

            # print(f"Invalid invoice format: {current_invoice}")
            return "Invalid Invoice"

        invoice_no = int(parts[0]) + 1
        fiscal_year = parts[1]
        today = datetime.today()
        month = today.month
        year = today.year

        # Determine the current fiscal year based on today's month
        if month > 3:  # After March, the fiscal year is the current year to next year
            fiscal_year_today = (
                f"{year % 100}-{(year + 1) % 100}"  # Example: 2024-25 will be 24-25
            )
        else:  # Before April, the fiscal year is last year to this year
            fiscal_year_today = (
                f"{(year - 1) % 100}-{year % 100}"  # Example: 2023-24 will be 23-24
            )

        # Check if the fiscal year in the current invoice matches the current fiscal year
        if fiscal_year != fiscal_year_today:
            # If it's not the current fiscal year, increment both the invoice number and fiscal year
            fiscal_year = fiscal_year_today
            invoice_no = 1  # Start from invoice number 1 for the new fiscal year

        # Return the new invoice number and fiscal year in the correct format
        return f"{invoice_no:02}/{fiscal_year}"

    except Exception as e:
        extras.print_colored(
            f"Invoice Number Error, Error generating invoice number: {str(e)}"
        )

        extras.print_colored(
            f"Invoice Number Error, Error generating invoice number: {str(e)}", "red"
        )

        # print("Invoice Number Error", f"Error generating invoice number: {str(e)}")
        return "Invalid Invoice"


# --------------------------------------------------------------------------------------------------------------------------------------

@app.route('/api/employees', methods=['GET'])
def get_employees():
    """Get all active employees from contract_user_info table"""
    try:
        # Query to get all active employees
        employees = db.session.query(
            contract_user_info.employee_name,
            contract_user_info.contractuserinfo_uid
        ).filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).all()
        
        # Format the result
        result = [
            {
                "employee_name": employee.employee_name,
                "user_id": employee.contractuserinfo_uid
            } for employee in employees
        ]
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/api/download-template', methods=['GET'])
def download_template():
    """Generate and download Excel template with predefined employee names and month/year"""
    month = request.args.get('month', '')
    year = request.args.get('year', '')
    
    if not month or not year:
        return jsonify({"error": "Both month and year parameters are required"}), 400
    
    try:
        # Query to get all active employees
        # Query to get all active employees, sorted alphabetically by name
        employees = db.session.query(
            contract_user_info.employee_name,
            contract_user_info.contractuserinfo_uid
        ).filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).order_by(
            contract_user_info.employee_name.asc()  # Sort alphabetically
        ).all()

        
        # Create DataFrame for Excel template
        data = []
        for i, employee in enumerate(employees, 1):
            month_year = f"{month} {year}"
            data.append({
                "Yes/No": "",
                "Serial No": i,
                "Name": employee.employee_name,
                "Month": month_year,
                "Payable Days": "",
                "Food Amount": "",
                "Arrears Month": "",
                "Arrears Payable Days": "",
                "Total Amount": ""
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Invoice Template', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Invoice Template']
            
            # Column setup
            arrears_month_col = 6  # "Arrears Month" column index (0-based)
            first_data_row = 1  # First data row (after header)
            last_data_row = len(data)

            # Step 1: Generate Month-Year list (current month and 11 months back)
            from datetime import datetime
            from dateutil.relativedelta import relativedelta
            import xlsxwriter

            # Parse the current month and year from parameters
            try:
                month_name = month.capitalize()
                month_number = datetime.strptime(month_name, '%B').month
                year_number = int(year)
                current_date = datetime(year_number, month_number, 1)
            except ValueError:
                return jsonify({"error": "Invalid month or year format"}), 400

            # Generate list of months - current month and 11 months back
            month_year_options = []
            # Add current month first
            month_year_options.append(current_date.strftime('%B %Y'))
            
            # Add previous 11 months
            for i in range(1, 12):
                previous_month = current_date - relativedelta(months=i)
                month_year_options.append(previous_month.strftime('%B %Y'))

            # Step 2: Add a hidden sheet for dropdown options
            dropdown_sheet = workbook.add_worksheet('DropdownData')
            dropdown_sheet.hide()

            for idx, option in enumerate(month_year_options):
                dropdown_sheet.write(idx, 0, option)

            # Step 3: Define a named range
            dropdown_range = f'DropdownData!$A$1:$A${len(month_year_options)}'
            workbook.define_name('MonthYearList', dropdown_range)

            # Step 4: Apply data validation using named range
            for row in range(first_data_row, last_data_row + 1):  # +1 to skip header
                cell = f'{xlsxwriter.utility.xl_col_to_name(arrears_month_col)}{row + 1}'
                worksheet.data_validation(cell, {
                    'validate': 'list',
                    'source': '=MonthYearList',
                    'input_title': 'Select Arrears Month',
                    'input_message': 'Select from current month and previous 11 months',
                })

            # Add a comment to the header cell
            worksheet.write_comment(0, arrears_month_col, 
                                    'This dropdown shows the current month and previous 11 months only.')

            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, column_width)

        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Invoice_Template_{month}_{year}.xlsx'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/upload-invoice-data', methods=['POST'])
def upload_invoice_data():
    """Process, validate, and calculate total salary for uploaded invoice data"""
    data = request.json
    
    if not data or not isinstance(data, list):
        return jsonify({"success": False, "message": "No valid data received"}), 400
    
    try:
        # Extract month and year from the first record's Month field (e.g., "January 2024")
        if not data or 'Month' not in data[0]:
            return jsonify({"success": False, "message": "No valid month data in records"}), 400
            
        month_year_parts = data[0]['Month'].strip().split()
        if len(month_year_parts) != 2:
            return jsonify({"success": False, "message": "Invalid month format in data"}), 400
            
        month = month_year_parts[0]
        year = month_year_parts[1]
        
        # Calculate days in month
        import calendar
        from datetime import datetime
        
        month_idx = datetime.strptime(month, '%B').month
        year_num = int(year)
        days_in_month = calendar.monthrange(year_num, month_idx)[1]
        
        # Get all valid employees from database with their salary information
        employees = db.session.query(
            contract_user_info.employee_name,
            contract_user_info.monthly_salary,
            contract_user_info.food_allowance_per_day_amount,
            contract_user_info.contractuserinfo_uid
        ).filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).all()
        
        # Create employee lookup dictionary with salary information
        employee_info = {}
        for emp in employees:
            employee_info[emp.employee_name.lower().strip()] = {
                'monthly_salary': float(emp.monthly_salary),
                'food_allowance_per_day': float(emp.food_allowance_per_day_amount),
                'employee_uid': emp.contractuserinfo_uid
            }
        
        valid_employee_names = set(employee_info.keys())
        
        # Validate each record and calculate salary
        validation_errors = []
        valid_records = []
        
        for idx, record in enumerate(data):
            row_num = idx + 1
            is_valid = True
            error_messages = []
            
            # Check if invoice is marked as "No" (rejected)
            yes_no_value = str(record.get('Yes/No', '')).strip().lower()
            if yes_no_value == 'no':
                # Skip validation for rejected invoices
                continue
            
            # 1. Validate Name (Required)
            emp_name_lower = str(record.get('Name', '')).lower().strip()
            if 'Name' not in record or not record['Name'] or emp_name_lower not in valid_employee_names:
                error_messages.append(f"Row {row_num}: Employee name not found in database")
                is_valid = False
            
            # 2. Validate Month (Required)
            expected_month_year = f"{month} {year}"
            if 'Month' not in record or not record['Month'] or str(record['Month']).strip() != expected_month_year:
                error_messages.append(f"Row {row_num}: Month does not match expected month/year ({expected_month_year})")
                is_valid = False
            
            # 3. Validate Payable Days (Required)
            try:
                payable_days = record.get('Payable Days', '')
                if payable_days == '':
                    error_messages.append(f"Row {row_num}: Payable Days is required")
                    is_valid = False
                else:
                    payable_days = int(payable_days)
                    if payable_days < 0 or payable_days > days_in_month:
                        error_messages.append(f"Row {row_num}: Payable Days must be between 0 and {days_in_month}")
                        is_valid = False
            except (ValueError, TypeError):
                error_messages.append(f"Row {row_num}: Payable Days must be a valid number")
                is_valid = False
            
            # 4. Validate Food Amount (Optional)
            food_amount = 0
            if 'Food Amount' in record and record['Food Amount'] and str(record['Food Amount']).strip() != "":
                try:
                    food_amount = float(record['Food Amount'])
                    if food_amount != int(food_amount):
                        error_messages.append(f"Row {row_num}: Food Amount must be a whole number")
                        is_valid = False
                except (ValueError, TypeError):
                    error_messages.append(f"Row {row_num}: Food Amount must be a valid number")
                    is_valid = False
            
            # 5. Validate Arrears Month format (Optional)
            arrears_month_days = 30  # Default value
            arrears_payable_days = 0
            if 'Arrears Month' in record and record['Arrears Month'] and str(record['Arrears Month']).strip() != "":
                import re
                pattern = r'^(January|February|March|April|May|June|July|August|September|October|November|December) \d{4}$'
                if not re.match(pattern, str(record['Arrears Month'])):
                    error_messages.append(f"Row {row_num}: Arrears Month is not in valid format")
                    is_valid = False
                else:
                    # 6. Validate Arrears Payable Days against arrears month days
                    try:
                        arr_month, arr_year = str(record['Arrears Month']).split()
                        arr_month_idx = datetime.strptime(arr_month, '%B').month
                        arr_year_num = int(arr_year)
                        arrears_month_days = calendar.monthrange(arr_year_num, arr_month_idx)[1]
                        
                        if 'Arrears Payable Days' in record and record['Arrears Payable Days'] and str(record['Arrears Payable Days']).strip() != "":
                            arrears_payable_days = int(record['Arrears Payable Days'])
                            if arrears_payable_days < 0 or arrears_payable_days > arrears_month_days:
                                error_messages.append(f"Row {row_num}: Arrears Payable Days must be between 0 and {arrears_month_days}")
                                is_valid = False
                    except (ValueError, TypeError):
                        error_messages.append(f"Row {row_num}: Arrears Payable Days must be a valid number")
                        is_valid = False
            elif 'Arrears Payable Days' in record and record['Arrears Payable Days'] and str(record['Arrears Payable Days']).strip() != "":
                error_messages.append(f"Row {row_num}: Arrears Payable Days provided but no Arrears Month specified")
                is_valid = False
            
            # Calculate total amount if record is valid
            if is_valid and emp_name_lower in employee_info:
                # Get employee salary information
                monthly_salary = employee_info[emp_name_lower]['monthly_salary']
                food_allowance_per_day = employee_info[emp_name_lower]['food_allowance_per_day']
                
                # Calculate main salary component
                main_payable_days_part = (monthly_salary / days_in_month) * payable_days
                
                # Calculate arrears component
                arrears_payable_days_part = (monthly_salary / arrears_month_days) * arrears_payable_days
                
                # Get food amount or calculate if not provided
                if food_amount == 0 and payable_days > 0:
                    food_amount = food_allowance_per_day * payable_days
                
                # Calculate total salary
                total_salary = main_payable_days_part + arrears_payable_days_part + food_amount
                
                # Round the values
                rounded_arrears_part = round(arrears_payable_days_part)
                rounded_main_part = round(main_payable_days_part)
                rounded_total = round(total_salary)
                
                # Add calculated values to record
                record['Total Amount'] = rounded_total
                record['Main Salary Component'] = rounded_main_part
                record['Arrears Component'] = rounded_arrears_part
                
                valid_records.append(record)
            else:
                for msg in error_messages:
                    validation_errors.append(msg)
        
        if not valid_records:
            return jsonify({
                "success": False, 
                "message": "No valid records to process",
                "errors": ["All records were rejected or contained validation errors"]
            }), 400
        
        if validation_errors:
            return jsonify({
                "success": False, 
                "message": "Validation errors found",
                "errors": validation_errors
            }), 400
        
        # Return success with the validated records (including calculated total amounts)
        return jsonify({
            "success": True, 
            "message": f"All {len(valid_records)} records validated successfully",
            "data": valid_records
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error processing data: {str(e)}"}), 500
    

from flask import request, jsonify
from datetime import datetime
import pytz
import calendar
from sqlalchemy import Column, Integer, String, Float, DateTime, Boolean

@app.route('/api/submit-invoice-data', methods=['POST'])
def submit_invoice_data():
    """Save validated invoice data with calculated total amount to database and generate templates"""
    try:
        # Add timeout and better error handling
        data = request.json
        
        if not data or not isinstance(data, list):
            return jsonify({"success": False, "message": "No valid data received"}), 400
        
        if len(data) == 0:
            return jsonify({"success": False, "message": "No data to process"}), 400
        
        # Get current location coordinates from request
        current_coordinates = request.headers.get("current-location", "")
        
        print(f"Processing {len(data)} invoice records...")
        
        # Validate data structure early
        if 'Month' not in data[0]:
            return jsonify({"success": False, "message": "Missing Month field in data"}), 400
        
        # Set batch size for incremental commits to prevent timeout
        BATCH_SIZE = 20  # Process 20 records at a time
        # Extract month and year from the first record
        month_year_parts = data[0]['Month'].strip().split()
        if len(month_year_parts) != 2:
            return jsonify({"success": False, "message": "Invalid month format"}), 400
            
        month_name = month_year_parts[0]
        year = int(month_year_parts[1])  # Convert year to int for calculations
        month_date = f"{month_name} {year}"
        
        # Convert month name to number (01-12)
        try:
            month_num = datetime.strptime(month_name, '%B').month
            year_num = int(year)
            days_in_month = calendar.monthrange(year_num, month_num)[1]
        except ValueError as e:
            return jsonify({"success": False, "message": f"Invalid month name: {month_name}"}), 400
        
        # Format month date in YYYY-MM format for database storage
        yyyy_mm_format = f"{year}-{month_num:02d}"
        
        # Create invoice date for template processing (use first day of month)
        invoice_date = f"{year}-{month_num:02d}-{days_in_month:02d}"
        
        # Determine fiscal year based on month
        # For months January(1), February(2), March(3): fiscal year is (year-1)-year
        # For months April(4) through December(12): fiscal year is year-(year+1)
        if 1 <= month_num <= 3:
            fiscal_year_start = year - 1
            fiscal_year_end = year
        else:  # 4 <= month_num <= 12
            fiscal_year_start = year
            fiscal_year_end = year + 1
        
        # Format the fiscal year part (e.g., "24-25")
        start_year_short = str(fiscal_year_start)[-2:]
        end_year_short = str(fiscal_year_end)[-2:]
        fiscal_year_format = f"{start_year_short}-{end_year_short}"
        
        # Define IST timezone for timestamp
        ist = pytz.timezone("Asia/Kolkata")
        current_time_ist = datetime.now(ist)
        
        # Get user info mapping (name to uid and other details) - optimize query
        print("Fetching employee data...")
        employee_map = {}
        employees = db.session.query(
            contract_user_info.contractuserinfo_uid,
            contract_user_info.employee_name,
            contract_user_info.manager_id,
            contract_user_info.template_no,
            contract_user_info.monthly_salary,
            contract_user_info.food_allowance_per_day_amount
        ).filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).all()
        
        for emp in employees:
            employee_map[emp.employee_name.lower().strip()] = {
                'uid': emp.contractuserinfo_uid,
                'manager_id': emp.manager_id,
                'template_no': emp.template_no,
                'monthly_salary': float(emp.monthly_salary),
                'food_allowance_per_day': float(emp.food_allowance_per_day_amount)
            }
        
        # Prepare for success/error tracking
        success_count = 0
        error_records = []
        processed_uids = []
        invoice_data_for_templates = []
        
        print(f"Starting to process {len(data)} records...")
        
        # Define the fiscal year range for database queries
        if 1 <= month_num <= 3:
            fiscal_start = f"{fiscal_year_start}-04-01"
            fiscal_end = f"{fiscal_year_end}-03-31"
        else:
            fiscal_start = f"{fiscal_year_start}-04-01"
            fiscal_end = f"{fiscal_year_end}-03-31"
        
        # Default location match to True if no coordinates provided
        location_match = True
        location_info = {}
        
        # Get manager_id for location verification
        # We'll use the manager_id from the first record's employee
        first_employee = None
        for record in data:
            if str(record.get('Yes/No', 'Yes')).strip().lower() == 'yes':
                emp_name = record['Name'].strip()
                emp_key = emp_name.lower()
                if emp_key in employee_map:
                    first_employee = employee_map[emp_key]
                    break
        
        # If we have a valid employee and coordinates are provided, verify location
        if first_employee and current_coordinates:
            manager_id = first_employee['manager_id']
            
            # Get manager's expected city coordinates
            manager_infos = manager_info.query.filter_by(managerinfo_uid=manager_id).first()
            if manager_infos and manager_infos.city:
                try:
                    # Get city name from current coordinates through reverse geocoding
                    current_city = get_city_from_coordinates(current_coordinates)
                    # Get expected city from manager's coordinates
                    expected_city = get_city_from_coordinates(manager_infos.city)
                    
                    location_match = (current_city.lower() == expected_city.lower())
                    location_info = {
                        "current_city": current_city,
                        "expected_city": expected_city,
                        "is_match": location_match
                    }
                    print("Current city:", current_city)
                    print("Expected city:", expected_city)
                    
                except Exception as e:
                    print(f"Error in location verification: {str(e)}")
                    # Continue with invoice generation even if location verification fails
                    pass
        
        # Prepare data for template processing
        invoice_data_for_templates = []
        
        processed_uids = []  # Track processed employee UIDs
        batch_counter = 0  # Counter for batch commits
        
        for i, record in enumerate(data):
            if i % 10 == 0:  # Log progress every 10 records
                print(f"Processing record {i+1}/{len(data)}...")
                
            # Get the Yes/No value (default to 'Yes' if not specified)
            yes_no_value = str(record.get('Yes/No', 'Yes')).strip()
            
            # Skip records marked as "No" but still process them to set the database value
            if yes_no_value.lower() == 'no':
                continue
                
            emp_name = record['Name'].strip()
            emp_key = emp_name.lower()
            
            if emp_key not in employee_map:
                error_records.append(f"Employee {emp_name} not found in database")
                continue
                
            employee_data = employee_map[emp_key]
            contract_employee_uid = employee_data['uid']
            manager_info_id = employee_data['manager_id']
            template_no = employee_data['template_no']
            monthly_salary = employee_data['monthly_salary']
            food_allowance_per_day = employee_data['food_allowance_per_day']
            
            # Process amounts and days
            try:
                # Use float for payable_days to preserve decimal values
                payable_days = float(record['Payable Days']) if record.get('Payable Days', '') else 0.0
                
                # Handle food amount - either use provided value or calculate based on payable days
                if record.get('Food Amount', '') and str(record.get('Food Amount', '')).strip():
                    food_amount = int(float(record['Food Amount']))  # Convert to int after float
                else:
                    food_amount = int(food_allowance_per_day * payable_days)  # Calculate and convert to int
                
                # Process arrears month (if exists) to YYYY-MM format
                arrears_month = record.get('Arrears Month', '')
                yyyy_mm_arrears = ''
                arrears_month_days = 30  # Default value
                
                if arrears_month:
                    arr_month, arr_year_str = arrears_month.strip().split()
                    arr_year = int(arr_year_str)
                    arr_month_num = datetime.strptime(arr_month, '%B').month
                    yyyy_mm_arrears = f"{arr_year}-{arr_month_num:02d}"
                    # Calculate days in arrears month
                    arrears_month_days = calendar.monthrange(arr_year, arr_month_num)[1]
                
                arrears_payable_days = int(record['Arrears Payable Days']) if record.get('Arrears Payable Days', '') else 0
                
                # Calculate monthly calculated salary and round to integer
                monthly_calculated_salary = int(round((monthly_salary / days_in_month) * payable_days))
                
                # Calculate arrears amount and round to integer
                arrears_amount = int(round((monthly_salary / arrears_month_days) * arrears_payable_days)) if arrears_payable_days > 0 else 0
                
                # Calculate total amount if not already provided
                if record.get('Total Amount', ''):
                    total_amount = int(float(record['Total Amount']))  # Convert to int after float
                else:
                    # Calculate total salary and round to integer
                    total_amount = int(round(monthly_calculated_salary + arrears_amount + food_amount))
                
            except ValueError as ve:
                error_records.append(f"Error converting values for {emp_name}: {str(ve)}")
                continue
            
            # Check if an entry already exists for this employee and month
            existing_invoice = db.session.query(invoice_no_table).filter(
                invoice_no_table.contract_employee_uid == str(contract_employee_uid),
                invoice_no_table.month_date == yyyy_mm_format,  # Use YYYY-MM format
                invoice_no_table.invoicenotable_isactive == 1
            ).first()
            
            # Find the maximum invoice number for this employee in the current fiscal year
            latest_invoice = db.session.query(invoice_no_table).filter(
                invoice_no_table.contract_employee_uid == str(contract_employee_uid),
                invoice_no_table.month_date >= fiscal_start,
                invoice_no_table.month_date <= fiscal_end,
                invoice_no_table.invoicenotable_isactive == 1
            ).order_by(invoice_no_table.last_date_generated.desc()).first()
            
            # Generate the invoice number
            if latest_invoice and latest_invoice.last_invoice_number:
                parts = latest_invoice.last_invoice_number.split('/')
                if len(parts) == 2 and parts[1] == fiscal_year_format:
                    # Same fiscal year, increment the number
                    try:
                        invoice_num = int(parts[0])
                        if existing_invoice and existing_invoice.last_invoice_number:
                            # If we're updating an existing invoice, use its number
                            invoice_number_format = existing_invoice.last_invoice_number
                        else:
                            # Otherwise, increment for a new invoice
                            next_num = invoice_num + 1
                            invoice_number_format = f"{next_num:02d}/{fiscal_year_format}"
                    except:
                        invoice_number_format = f"01/{fiscal_year_format}"
                else:
                    # Different fiscal year, start from 01
                    invoice_number_format = f"01/{fiscal_year_format}"
            else:
                # No invoices in this fiscal year, start from 01
                invoice_number_format = f"01/{fiscal_year_format}"
            
            if existing_invoice:
                # Update existing record - use explicit column assignment
                existing_invoice.payable_days = payable_days
                existing_invoice.food_amount = food_amount
                existing_invoice.arrears_month = yyyy_mm_arrears
                existing_invoice.arrears_payable_days = arrears_payable_days
                existing_invoice.total_amount = total_amount
                existing_invoice.last_date_generated = current_time_ist
                existing_invoice.yes_no = yes_no_value
                
                # Update coordinates if provided
                if current_coordinates:
                    existing_invoice.coordinates = current_coordinates
                
                # Add to history with location_match flag inverted
                add_history(user_uid=contract_employee_uid, date=yyyy_mm_format, 
                           invoice_no=existing_invoice.last_invoice_number, wrong_location=not location_match)
                
                current_invoice_no = existing_invoice.last_invoice_number
            else:
                # Create new invoice record
                new_invoice = invoice_no_table()
                new_invoice.last_invoice_number = invoice_number_format
                new_invoice.month_date = yyyy_mm_format
                new_invoice.payable_days = payable_days
                new_invoice.food_amount = food_amount
                new_invoice.arrears_month = yyyy_mm_arrears
                new_invoice.arrears_payable_days = arrears_payable_days
                new_invoice.total_amount = total_amount
                new_invoice.contract_employee_uid = str(contract_employee_uid)
                new_invoice.manager_info_id = manager_info_id
                new_invoice.invoicenotable_isactive = 1
                new_invoice.last_date_generated = current_time_ist
                new_invoice.yes_no = yes_no_value
                
                # Store coordinates if provided
                if current_coordinates:
                    new_invoice.coordinates = current_coordinates
                
                db.session.add(new_invoice)
                # Add to history with location_match flag inverted
                add_history(user_uid=contract_employee_uid, date=yyyy_mm_format, 
                           invoice_no=invoice_number_format, wrong_location=not location_match)
                
                current_invoice_no = invoice_number_format
            processed_uids.append(str(contract_employee_uid))
            # Prepare data for template processing with integer values
            invoice_data_for_templates.append({
                "uid": contract_employee_uid,
                "invoice_no": current_invoice_no,
                "payable_days": payable_days,
                "monthly_calculated_salary": str(monthly_calculated_salary),  # Convert to string to avoid float
                "arrears_month": yyyy_mm_arrears,
                "arrears_payable_days": arrears_payable_days,
                "arrears_amount": str(arrears_amount),  # Convert to string to avoid float
                "food_amount_salary": str(food_amount),  # Convert to string to avoid float
                "total_calculated_salary": str(total_amount),  # Convert to string to avoid float
                "amount_in_words": ""  # Will be calculated in process_templates
            })
            
            success_count += 1
            batch_counter += 1
            
            # Commit in batches to prevent connection timeout
            if batch_counter >= BATCH_SIZE:
                try:
                    db.session.commit()
                    print(f"Batch commit successful: {success_count} records processed so far")
                    batch_counter = 0  # Reset counter
                except Exception as commit_error:
                    print(f"Error during batch commit: {str(commit_error)}")
                    db.session.rollback()
                    raise
        
        # Commit any remaining records
        if batch_counter > 0:
            db.session.commit()
            print(f"Final commit successful: {success_count} total records processed")
        
        # Generate templates if we have valid invoice data
        output_folder = None
        manager_zip_files = {}  # Store zip files per manager
        managers_data = {}
        template_generation_success = False
        
        # ALWAYS generate templates - this is critical for manager workflow
        # Template generation happens after data is committed to database
        if invoice_data_for_templates:
            try:
                print(f"Starting template generation for {len(invoice_data_for_templates)} records...")
                
                # Group invoice data by manager (optimize with single query)
                manager_groups = {}
                
                # Fetch all employees in one query instead of individual queries
                uids_to_fetch = [inv["uid"] for inv in invoice_data_for_templates]
                employees_batch = db.session.query(
                    contract_user_info.contractuserinfo_uid,
                    contract_user_info.manager_id
                ).filter(
                    contract_user_info.contractuserinfo_uid.in_(uids_to_fetch)
                ).all()
                
                # Create a map for quick lookup
                uid_to_manager = {emp.contractuserinfo_uid: emp.manager_id for emp in employees_batch}
                
                # Group by manager
                for invoice_row in invoice_data_for_templates:
                    uid = invoice_row["uid"]
                    manager_id = uid_to_manager.get(uid)
                    
                    if manager_id:
                        if manager_id not in manager_groups:
                            manager_groups[manager_id] = []
                        manager_groups[manager_id].append(invoice_row)
                
                # Process templates for each manager separately with timeout protection
                total_managers = len(manager_groups)
                current_manager_count = 0
                
                for manager_id, manager_invoice_data in manager_groups.items():
                    try:
                        current_manager_count += 1
                        print(f"[{current_manager_count}/{total_managers}] Generating templates for manager {manager_id} ({len(manager_invoice_data)} records)...")
                        
                        # Process templates for this manager's employees only
                        output_folder = process_templates_bulk(manager_invoice_data, invoice_date, 
                                                             current_coordinates, location_match)
                        
                        # Create a unique zip file for this manager
                        print(f"[{current_manager_count}/{total_managers}] Creating ZIP file for manager {manager_id}...")
                        manager_zip_path = extras.zip_folder(output_folder, f"{yyyy_mm_format}_manager_{manager_id}")
                        manager_zip_files[manager_id] = manager_zip_path
                        
                        print(f"✓ Manager {manager_id} complete: {manager_zip_path}")
                        
                    except Exception as e:
                        print(f"✗ Error generating templates for manager {manager_id}: {str(e)}")
                        import traceback
                        traceback.print_exc()
                        # Continue with other managers even if one fails
                        continue
                
                # Get managers data for the modal
                if manager_zip_files:
                    managers_data = get_managers_for_employees(invoice_data_for_templates)
                    template_generation_success = True
                
            except Exception as e:
                print(f"Error generating templates: {str(e)}")
                import traceback
                traceback.print_exc()
                # Even if template generation fails, still return success for data save
                # but indicate template failure
                template_generation_success = False
        
        print(f"Successfully processed {success_count} records")
        
        # Prepare response message
        if template_generation_success:
            message = f"Successfully processed {success_count} invoice records and generated templates"
        elif len(manager_zip_files) > 0:
            message = f"Successfully processed {success_count} invoice records and generated templates (with some warnings)"
        else:
            message = f"Successfully processed {success_count} invoice records. Template generation failed - please retry or use automation feature"
        
        return jsonify({
            "success": True,
            "message": message,
            "errors": error_records if error_records else None,
            "location_info": location_info if location_info else None,
            "templates_generated": len(manager_zip_files) > 0,
            "manager_zip_files": manager_zip_files,
            "managers_data": managers_data,
            "processed_uids": processed_uids
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error in submit_invoice_data: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error processing data: {str(e)}"}), 500


def process_templates_bulk(invoice_data, invoice_date, current_coordinates="", location_match=True):
    """Process templates for bulk uploaded invoice data"""
    print("Starting process_templates_bulk function")
    print("Invoice Date:", invoice_date)
    print("Invoice Data Count:", len(invoice_data))
    print("Current Coordinates:", current_coordinates)
    print("Location Match:", location_match)

    # Extract UIDs from invoice data
    uids = [row["uid"] for row in invoice_data]
    print("UIDs:", uids)

    # Query database for matching rows
    matching_users = contract_user_info.query.filter(
        contract_user_info.contractuserinfo_uid.in_(uids)
    ).all()
    print("Matching users count:", len(matching_users))
    
    # Ensure at least one user is found
    if not matching_users:
        raise Exception("No matching users found in database")

    # Extract manager_id from the first matching user
    manager_id = matching_users[0].manager_id
    print("Manager ID:", manager_id)
    
    # Create a dictionary to map UID to the database row
    db_data_map = {str(user.contractuserinfo_uid): user for user in matching_users}
    print("DB data map created")
    
    # Prepare final data to send to process_templates
    matched_data = []
    for row in invoice_data:
        uid = str(row["uid"])  # Ensure string type for consistency
        if uid in db_data_map:
            matched_data.append({
                "db_row": db_data_map[uid].to_dict(),  # Convert DB row to dictionary
                "table_row": row,  # Invoice row from request
            })

    # Pass manager_id and location info to process_templates (reuse existing function)
    output_folder = process_templates(matched_data, invoice_date, manager_id, 
                                    current_coordinates, location_match)
    print("Output folder:", output_folder)
    
    return output_folder


def get_managers_for_employees(invoice_data):
    """Get unique managers for the processed employees"""
    try:
        # Extract UIDs from invoice data
        uids = [row["uid"] for row in invoice_data]
        
        # Query to get unique managers for these employees
        managers_query = db.session.query(
            manager_info.managerinfo_uid,
            manager_info.name,
            manager_info.manager_email
        ).join(
            contract_user_info, 
            contract_user_info.manager_id == manager_info.managerinfo_uid
        ).filter(
            contract_user_info.contractuserinfo_uid.in_(uids),
            manager_info.isactive == 1
        ).distinct().all()
        
        managers_data = []
        for manager in managers_query:
            managers_data.append({
                "manager_id": manager.managerinfo_uid,
                "manager_name": manager.name,
                "manager_email": manager.manager_email
            })
        
        return managers_data
    except Exception as e:
        print(f"Error getting managers data: {str(e)}")
        return []


import traceback  # Add this at the top if not already imported
@app.route('/api/send-templates-to-managers', methods=['POST'])
def send_templates_to_managers():
    """Send generated templates to selected managers via email"""
    try:
        data = request.json
        print("Received data:", data)

        selected_managers = data.get('selected_managers', [])
        manager_zip_files = data.get('manager_zip_files', {})  # Changed from single zip_file_path
        month_year = data.get('month_year', '')
        processed_uids = data.get('processed_uids', [])

        print(f"Selected managers: {selected_managers}")
        print(f"Manager zip files: {manager_zip_files}")
        print(f"Month-Year: {month_year}")
        print(f"Processed UIDs received: {processed_uids}")

        if not selected_managers:
            print("No managers selected.")
            return jsonify({"success": False, "message": "No managers selected"}), 400

        if not manager_zip_files:
            print("No zip files found.")
            return jsonify({"success": False, "message": "Template files not found"}), 400

        success_count = 0
        error_count = 0

        for manager_id in selected_managers:
            try:
                print(f"Processing manager ID: {manager_id}")

                # Get manager details
                manager = manager_info.query.filter_by(managerinfo_uid=manager_id).first()
                print(f"Manager object: {manager}")

                if not manager or not manager.manager_email:
                    print(f"Manager {manager_id} not found or missing email.")
                    error_count += 1
                    continue

                # Get the specific zip file for this manager
                manager_zip_path = manager_zip_files.get(str(manager_id))
                if not manager_zip_path or not os.path.exists(manager_zip_path):
                    print(f"Zip file not found for manager {manager_id}: {manager_zip_path}")
                    error_count += 1
                    continue

                # Get invoice data for this manager's employees, filtered by processed UIDs
                invoice_table_data = get_invoice_data_for_manager(manager_id, month_year, processed_uids)

                # Log invoice data for debugging
                print(f"Invoice data for manager {manager_id}: {[row['name'] for row in invoice_table_data]}")

                # Prepare email
                subject = f"Invoice Templates - {month_year}"
                
                # Create HTML email body with table
                body_html = create_invoice_email_html(manager.name, month_year, invoice_table_data)

                print(f"Sending email to {manager.manager_email} with attachment {manager_zip_path}")

                # Send email with manager-specific zip file
                send_email_with_attachment_html(
                    to_email=manager.manager_email,
                    subject=subject,
                    body_html=body_html,
                    attachment_path=manager_zip_path
                )

                print(f"Email sent to {manager.manager_email}")
                success_count += 1

            except Exception as e:
                print(f"Error sending email to manager {manager_id}: {str(e)}")
                traceback.print_exc()
                error_count += 1

        return jsonify({
            "success": True,
            "message": f"Templates sent to {success_count} managers successfully",
            "success_count": success_count,
            "error_count": error_count
        })

    except Exception as e:
        print("An unexpected error occurred in the outer try block.")
        print(f"Exception: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error sending templates: {str(e)}"}), 500




def get_invoice_data_for_manager(manager_id, month_year, processed_uids=None):
    """Get invoice data for employees under a specific manager, filtered by processed UIDs"""
    try:
        from datetime import datetime
        import calendar
        
        # Parse month_year to get YYYY-MM format
        if month_year:
            month_name, year = month_year.strip().split()
            month_num = datetime.strptime(month_name, '%B').month
            yyyy_mm_format = f"{year}-{month_num:02d}"
        else:
            now = datetime.now()
            yyyy_mm_format = f"{now.year}-{now.month:02d}"

        # Log input parameters for debugging
        print(f"Fetching invoice data for manager_id: {manager_id}, month_year: {month_year}, processed_uids: {processed_uids}")

        # Query to get invoice data for employees under this manager
        query = db.session.query(
            contract_user_info.employee_name,
            invoice_no_table.last_invoice_number,
            invoice_no_table.payable_days,
            invoice_no_table.food_amount,
            invoice_no_table.total_amount,
            projects_table.project_name
        ).join(
            invoice_no_table,
            invoice_no_table.contract_employee_uid == contract_user_info.contractuserinfo_uid
        ).outerjoin(
            projects_table,
            projects_table.projectstable_uid == invoice_no_table.project_uid
        ).filter(
            contract_user_info.manager_id == manager_id,
            invoice_no_table.month_date == yyyy_mm_format,
            invoice_no_table.invoicenotable_isactive == 1,
            contract_user_info.contractuserinfo_isactive == 1
        )

        # Apply UID filter if provided
        if processed_uids:
            # Ensure UIDs are strings to match database format
            processed_uids = [str(uid) for uid in processed_uids]
            print(f"Applying UID filter with: {processed_uids}")
            query = query.filter(invoice_no_table.contract_employee_uid.in_(processed_uids))
        else:
            print("No processed_uids provided, fetching all active invoices for manager")

        # Execute query and log results
        invoice_data = query.all()
        print(f"Query returned {len(invoice_data)} records: {[row.employee_name for row in invoice_data]}")

        # Calculate monthly calculated salary for each employee
        result_data = []
        for row in invoice_data:
            # Get employee's monthly salary and days in month
            employee = contract_user_info.query.filter_by(
                employee_name=row.employee_name,
                contractuserinfo_isactive=1
            ).first()
            
            if employee:
                # Calculate days in the month
                year, month = yyyy_mm_format.split('-')
                days_in_month = calendar.monthrange(int(year), int(month))[1]
                
                # Calculate monthly calculated salary
                monthly_calculated_salary = (float(employee.monthly_salary) / days_in_month) * float(row.payable_days)
                
                result_data.append({
                    'name': row.employee_name,
                    'invoice_no': row.last_invoice_number,
                    'payable_days': row.payable_days,
                    'food_amount': round(row.food_amount) if row.food_amount else 0,
                    'monthly_calculated_salary': round(monthly_calculated_salary, 2),
                    'total_calculated_salary': round(row.total_amount) if row.total_amount else 0,
                    'project': row.project_name if row.project_name else 'N/A'
                })

        print(f"Returning {len(result_data)} processed records")
        return result_data

    except Exception as e:
        print(f"Error getting invoice data for manager {manager_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def create_invoice_email_html(manager_name, month_year, invoice_data):
    """Create HTML email body with invoice data table"""
    
    # Start building HTML
    html = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333;
                background-color: #f8f9fa;
                margin: 0;
                padding: 20px;
            }}
            .container {{
                max-width: 1000px;
                margin: 0 auto;
                background-color: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            .header {{
                text-align: center;
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 3px solid #007bff;
            }}
            .header h1 {{
                color: #007bff;
                margin: 0;
                font-size: 28px;
            }}
            .header p {{
                color: #666;
                margin: 10px 0 0 0;
                font-size: 16px;
            }}
            .greeting {{
                margin-bottom: 25px;
                font-size: 16px;
            }}
            .table-container {{
                margin: 20px 0;
                overflow-x: auto;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
                background-color: white;
                box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                border-radius: 8px;
                overflow: hidden;
            }}
            th {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                font-weight: 600;
                padding: 15px 12px;
                text-align: left;
                font-size: 14px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            td {{
                padding: 12px;
                border-bottom: 1px solid #e9ecef;
                font-size: 14px;
            }}
            tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            tr:hover {{
                background-color: #e3f2fd;
                transition: background-color 0.3s ease;
            }}
            .amount {{
                text-align: right;
                font-weight: 600;
                color: #28a745;
            }}
            .invoice-no {{
                font-family: 'Courier New', monospace;
                background-color: #e9ecef;
                padding: 4px 8px;
                border-radius: 4px;
                font-weight: 600;
            }}
            .total-row {{
                background-color: #f1f3f4 !important;
                font-weight: 600;
                border-top: 2px solid #007bff;
            }}
            .footer {{
                margin-top: 30px;
                padding-top: 20px;
                border-top: 1px solid #e9ecef;
                font-size: 14px;
                color: #666;
            }}
            .attachment-note {{
                background-color: #fff3cd;
                border: 1px solid #ffeaa7;
                border-radius: 5px;
                padding: 15px;
                margin: 20px 0;
                font-size: 14px;
            }}
            .attachment-note strong {{
                color: #856404;
            }}
            @media (max-width: 600px) {{
                .container {{
                    padding: 15px;
                }}
                table {{
                    font-size: 12px;
                }}
                th, td {{
                    padding: 8px 6px;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Invoice Templates</h1>
                <p>{month_year}</p>
            </div>
            
            <div class="greeting">
                <p>Dear {manager_name},</p>
                <p>Please find below the invoice details for <strong>{month_year}</strong> and the attached template files.</p>
            </div>
    """

    if invoice_data:
        # Add the invoice data table
        html += """
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Name</th>
                            <th>Invoice No</th>
                            <th>Payable Days</th>
                            <th>Food Amount Salary</th>
                            <th>Monthly Calculated Salary</th>
                            <th>Total Calculated Salary</th>
                            <th>Project</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        # Add data rows
        total_food_amount = 0
        total_monthly_salary = 0
        total_calculated_salary = 0
        
        for row in invoice_data:
            total_food_amount += row['food_amount']
            total_monthly_salary += row['monthly_calculated_salary']
            total_calculated_salary += row['total_calculated_salary']
            
            html += f"""
                        <tr>
                            <td>{row['name']}</td>
                            <td><span class="invoice-no">{row['invoice_no']}</span></td>
                            <td style="text-align: center;">{row['payable_days']}</td>
                            <td class="amount">₹{row['food_amount']:,.2f}</td>
                            <td class="amount">₹{row['monthly_calculated_salary']:,.2f}</td>
                            <td class="amount">₹{row['total_calculated_salary']:,.2f}</td>
                            <td>{row['project']}</td>
                        </tr>
            """
        
        
        html += """
                    </tbody>
                </table>
            </div>
        """
    else:
        html += """
            <div style="text-align: center; padding: 40px; background-color: #f8f9fa; border-radius: 8px; margin: 20px 0;">
                <p style="color: #6c757d; font-size: 16px;">No invoice data found for this period.</p>
            </div>
        """

    # Add attachment note and footer
    html += """
            <div class="attachment-note">
                <strong>Attachment:</strong> Please find the invoice template files attached as a ZIP archive.
            </div>
            
            <div class="footer">
                <p>Best regards,<br>
                <strong>Invoice Management System</strong></p>
                <hr style="border: none; border-top: 1px solid #e9ecef; margin: 15px 0;">
                <p style="font-size: 12px; color: #999;">
                    This is an automated email. Please do not reply to this message.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html


def send_email_with_attachment_html(to_email, subject, body_html, attachment_path):
    """Send HTML email with attachment - enhanced version with proper UTF-8 encoding"""
    
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import os
    
    # Email configuration (you should move these to environment variables)
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "shaikhazeem4646@gmail.com"
    sender_password = "kkgljpsvornmzois"
    
    try:
        # Create message with proper encoding policy
        msg = MIMEMultipart('alternative')
        msg['From'] = sender_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add HTML body to email with UTF-8 encoding
        html_part = MIMEText(body_html, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Open file in binary mode
        with open(attachment_path, "rb") as attachment:
            # Instance of MIMEBase and named as part
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        # Encode file in ASCII characters to send by email    
        encoders.encode_base64(part)
        
        # Add header as key/value pair to attachment part
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {os.path.basename(attachment_path)}',
        )
        
        # Attach the part to message
        msg.attach(part)
        
        # Create SMTP session
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable security
        server.login(sender_email, sender_password)
        
        # Convert message to bytes with proper UTF-8 encoding
        # This is the key fix - use as_bytes() instead of as_string()
        text = msg.as_bytes()
        server.send_message(msg)  # Use send_message() instead of sendmail() for better encoding support
        server.quit()
        
        print(f"HTML email sent successfully to {to_email}")
        
    except Exception as e:
        print(f"Error sending email to {to_email}: {str(e)}")
        raise e


# Alternative solution using the original approach with proper encoding
def send_email_with_attachment_html_alternative(to_email, subject, body_html, attachment_path):
    """Alternative approach - using sendmail with proper string encoding"""
    
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import os
    
    # Email configuration
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "shaikhazeem4646@gmail.com"
    sender_password = "kkgljpsvornmzois"
    
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['From'] = sender_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add HTML body to email with UTF-8 encoding
        html_part = MIMEText(body_html, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Open file in binary mode
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {os.path.basename(attachment_path)}',
        )
        
        msg.attach(part)
        
        # Create SMTP session
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        
        # The key difference: use as_string() but encode properly for sendmail
        text = msg.as_string()
        # Convert both emails to list format and encode the message
        server.sendmail(sender_email, [to_email], text.encode('utf-8'))
        server.quit()
        
        print(f"HTML email sent successfully to {to_email}")
        
    except Exception as e:
        print(f"Error sending email to {to_email}: {str(e)}")
        raise e


# Keep the original function as fallback
def send_email_with_attachment(to_email, subject, body, attachment_path):
    """Send plain text email with attachment - original version kept for compatibility"""
    
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import os
    
    # Email configuration (you should move these to environment variables)
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "shaikhazeem4646@gmail.com"
    sender_password = "kkgljpsvornmzois"
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add body to email
        msg.attach(MIMEText(body, 'plain'))
        
        # Open file in binary mode
        with open(attachment_path, "rb") as attachment:
            # Instance of MIMEBase and named as part
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        # Encode file in ASCII characters to send by email    
        encoders.encode_base64(part)
        
        # Add header as key/value pair to attachment part
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {os.path.basename(attachment_path)}',
        )
        
        # Attach the part to message
        msg.attach(part)
        
        # Create SMTP session
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable security
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, to_email, text)
        server.quit()
        
        print(f"Email sent successfully to {to_email}")
        
    except Exception as e:
        print(f"Error sending email to {to_email}: {str(e)}")

@app.route("/update_employee_project", methods=["POST"])
def update_employee_project():
    try:
        data = request.json
        employee_id = data.get("employee_id")
        project_id = data.get("project_id")
        
        if not employee_id or not project_id:
            return jsonify({"error": "Missing required fields"}), 400
        
        # Get the employee record
        employee = contract_user_info.query.get(employee_id)
        
        if not employee:
            return jsonify({"error": "Employee not found"}), 404
        
        # Update project ID
        employee.project_id = project_id
        db.session.commit()
        
        return jsonify({"success": True, "message": "Project updated successfully"}), 200
    
    except Exception as e:
        db.session.rollback()
        print(f"Error updating employee project: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/update_invoice_data', methods=['POST'])
def update_invoice_data():
    data = request.json
    
    try:
        # Fetch the employee data to get project_uid and manager_id
        # Fetch the employee data to get project_uid and manager_id
        employee = contract_user_info.query.filter_by(
        contractuserinfo_uid=data['contract_employee_uid']
        ).first()

        if not employee:
            return jsonify({'success': False, 'error': 'Employee not found'})

# Get project_uid from the employee record
        project_id = employee.project_id

# Use manager_info_id from input, or default to employee.manager_id
        manager_info_id = data.get('manager_info_id') or employee.manager_id

        
        invoice = invoice_no_table.query.filter_by(
            contract_employee_uid=data['contract_employee_uid'],
            month_date=data['month_date']
        ).first()
        
        updated = False  # Track if any update happened

        if not invoice:
            invoice = invoice_no_table(
    contract_employee_uid=data['contract_employee_uid'],
    month_date=data['month_date'],
    last_invoice_number=data['invoice_no'] if 'invoice_no' in data and data['invoice_no'] else '',
    invoicenotable_isactive=1,
    # Automatically add project_uid and manager_info_id
    project_uid=project_id,
    manager_info_id=manager_info_id,
    remark = data.get('remark', ''),  # This is looking for 'remark', not 'remarks'
)
            db.session.add(invoice)
            updated = True  # New record added, mark as updated
        else:
            # Always ensure project_uid and manager_info_id are set correctly on updates
            if invoice.project_uid != project_id:
                invoice.project_uid = project_id
                updated = True
                
            if invoice.manager_info_id != manager_info_id:
                invoice.manager_info_id = manager_info_id
                updated = True

        # Check and update fields
        if 'invoice_no' in data and data['invoice_no'] and invoice.last_invoice_number != data['invoice_no']:
            invoice.last_invoice_number = data['invoice_no']
            updated = True

        if 'payable_days' in data and data['payable_days'] and invoice.payable_days != data['payable_days']:
            invoice.payable_days = data['payable_days']
            updated = True

        if 'food_amount' in data:
            try:
                food_amt = float(data['food_amount']) if data['food_amount'] else 0
            except ValueError:
                food_amt = 0
            if invoice.food_amount != food_amt:
                invoice.food_amount = food_amt
                updated = True

        if 'remark' in data:
            invoice.remark = data['remark']
            updated = True

        if 'arrears_month' in data and invoice.arrears_month != data['arrears_month']:
            invoice.arrears_month = data['arrears_month']
            updated = True

        if 'arrears_payable_days' in data:
            try:
                arrears_days = int(data['arrears_payable_days']) if data['arrears_payable_days'] else 0
            except ValueError:
                arrears_days = 0
            if invoice.arrears_payable_days != arrears_days:
                invoice.arrears_payable_days = arrears_days
                updated = True

        if 'total_amount' in data and data['total_amount'] and invoice.total_amount != data['total_amount']:
            invoice.total_amount = data['total_amount']
            updated = True

        # If project_id is provided explicitly in the data, use it (overriding the default)
        if 'project_id' in data and data['project_id'] and invoice.project_uid != data['project_id']:
            invoice.project_uid = data['project_id']
            updated = True

        # Set 'yes_no' if updated and not already 'yes'
        if updated and invoice.yes_no != 'Yes':
            invoice.yes_no = 'Yes'

        db.session.commit()
        return jsonify({'success': True, 'message': 'Invoice updated successfully'})
    
    except Exception as e:
        db.session.rollback()
        print(f"Error updating invoice data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/sidebar', methods=['POST'])
def sidebar():
    """API endpoint to get report data based on date range"""
    try:
        data = request.get_json()
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        # Parse dates (format: MM/YYYY)
        start_date = datetime.strptime(start_date_str, '%m/%Y')
        end_date = datetime.strptime(end_date_str, '%m/%Y')
        
        # Convert to string format for database comparison (YYYY-MM)
        start_month_year = start_date.strftime('%Y-%m')
        end_month_year = end_date.strftime('%Y-%m')


        query = text("""
            SELECT 
                cu.employee_name,
                inv.payable_days,
                inv.food_amount,
                p.project_name,
                inv.last_invoice_number,
                (inv.payable_days * cu.monthly_salary / 30) AS total_payable,
                (inv.payable_days * cu.food_allowance_per_day_amount) AS total_food,
                inv.arrears_payable_days,
                inv.total_amount,
                m.name AS manager_name,
                inv.month_date
            FROM invoice_no_table inv
            LEFT JOIN contract_user_info cu ON inv.contract_employee_uid = CAST(cu.contractuserinfo_uid AS CHAR)
            LEFT JOIN projects_table p ON inv.project_uid = p.projectstable_uid
            LEFT JOIN manager_info m ON inv.manager_info_id = m.managerinfo_uid
            WHERE 
                inv.invoicenotable_isactive = 1
                AND inv.month_date >= :start_month_year
                AND inv.month_date <= :end_month_year
            ORDER BY cu.employee_name, inv.month_date
        """)
        result = db.session.execute(query, {
            "start_month_year": start_month_year,
            "end_month_year": end_month_year
        })
# Add 'month' to response
        data = []
        for row in result:
            data.append({
        "employee_name": row.employee_name,
        "payable_days": row.payable_days,
        "food_amount": float(row.food_amount or 0),
        "project_name": row.project_name,
        "last_invoice_number": row.last_invoice_number,
        "total_payable": float(row.total_payable or 0),
        "total_food": float(row.total_food or 0),
        "arrears_payable_days": row.arrears_payable_days,
        "total_amount": float(row.total_amount or 0),
        "manager_name": row.manager_name,
        "month": row.month_date  # 🆕 Add month here
    })


        
        return jsonify({"success": True, "data": data})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})



def parse_month_date(month_date_str):
    """Parse the month_date field to extract month and year"""
    if not month_date_str:
        return None, None

    formats = [
        r'([A-Za-z]+)\s+(\d{4})',      # "April 2024"
        r'(\d{1,2})/(\d{4})',          # "4/2024"
        r'(\d{4})-(\d{1,2})'           # "2024-04"
    ]

    for pattern in formats:
        match = re.match(pattern, month_date_str.strip())
        if match:
            part1, part2 = match.groups()
            try:
                if pattern == r'(\d{4})-(\d{1,2})':  # YYYY-MM
                    year = int(part1)
                    month = int(part2)
                elif part1.isalpha():
                    month_names = {
                        "jan": 1, "january": 1,
                        "feb": 2, "february": 2,
                        "mar": 3, "march": 3,
                        "apr": 4, "april": 4,
                        "may": 5,
                        "jun": 6, "june": 6,
                        "jul": 7, "july": 7,
                        "aug": 8, "august": 8,
                        "sep": 9, "september": 9,
                        "oct": 10, "october": 10,
                        "nov": 11, "november": 11,
                        "dec": 12, "december": 12
                    }
                    month = month_names.get(part1.lower())
                    year = int(part2)
                else:
                    month = int(part1)
                    year = int(part2)

                return month, year
            except ValueError:
                continue

    return None, None

@app.route('/api/report-data', methods=['POST'])
def get_report_data():
    """API endpoint to fetch report data based on date range"""
    data = request.json
    start_month = int(data.get('start_month'))
    start_year = int(data.get('start_year'))
    end_month = int(data.get('end_month'))
    end_year = int(data.get('end_year'))

    all_invoices = invoice_no_table.query.filter(
        invoice_no_table.invoicenotable_isactive == 1
    ).all()

    filtered_invoices = []
    monthly_totals = {}
    months_in_range = []

    # Generate all months in the date range
    current_year, current_month = start_year, start_month
    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
        month_key = f"{calendar.month_name[current_month]} {current_year}"
        months_in_range.append(month_key)
        monthly_totals[month_key] = 0
        
        # Move to next month
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

    for invoice in all_invoices:
        month, year = parse_month_date(invoice.month_date)
        if month is None or year is None:
            continue

        if (year > start_year or (year == start_year and month >= start_month)) and \
           (year < end_year or (year == end_year and month <= end_month)):
            filtered_invoices.append(invoice)

            month_key = f"{calendar.month_name[month]} {year}"
            if invoice.total_amount:
                monthly_totals[month_key] += float(invoice.total_amount)

    # Preload project names
    project_lookup = {
        p.projectstable_uid: p.project_name
        for p in projects_table.query.filter_by(isactive=1).all()
    }

    # Aggregate by project and month
    project_data = {}
    for invoice in filtered_invoices:
        pid = invoice.project_uid
        month, year = parse_month_date(invoice.month_date)
        month_key = f"{calendar.month_name[month]} {year}"
        
        if pid not in project_data:
            project_data[pid] = {
                'monthly_data': {}, 
                'total_count': 0, 
                'total_cost': 0
            }
            
        # Initialize month data if not exists
        if month_key not in project_data[pid]['monthly_data']:
            project_data[pid]['monthly_data'][month_key] = {'count': 0, 'cost': 0}
            
        # Update counts and costs
        project_data[pid]['monthly_data'][month_key]['count'] += 1
        if invoice.total_amount:
            cost = float(invoice.total_amount)
            project_data[pid]['monthly_data'][month_key]['cost'] += cost
            project_data[pid]['total_cost'] += cost
        project_data[pid]['total_count'] += 1

    # Format results
    formatted_results = []
    for pid, data in project_data.items():
        project_name = project_lookup.get(pid, f"Project {pid}")
        formatted_results.append({
            'project_name': project_name,
            'monthly_data': data['monthly_data'],
            'total_count': data['total_count'],
            'total_cost': data['total_cost']
        })

    # Prepare chart data
    chart_data = [{
        'month': month_key,
        'total_cost': monthly_totals[month_key]
    } for month_key in months_in_range]

    return jsonify({
        'table_data': formatted_results,
        'chart_data': chart_data,
        'months': months_in_range,
        'total_count': sum(item['total_count'] for item in formatted_results),
        'total_cost': sum(item['total_cost'] for item in formatted_results)
    })
@app.route('/api/employees-with-salary', methods=['GET'])
def get_employees_with_salary():
    """Return all active employees with their salary information"""
    try:
        employees = db.session.query(
            contract_user_info.employee_name,
            contract_user_info.monthly_salary,
            contract_user_info.food_allowance_per_day_amount
        ).filter(
            contract_user_info.contractuserinfo_isactive == 1
        ).all()
        
        result = []
        for emp in employees:
            result.append({
                'employee_name': emp.employee_name,
                'monthly_salary': float(emp.monthly_salary),
                'food_allowance_per_day_amount': float(emp.food_allowance_per_day_amount)
            })
        
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/send_remark_email', methods=['POST'])
def send_remark_email():
    data = request.json
    employee_name = data.get('employee_name')
    remark = data.get('remark')
    month = data.get('month')
    manager_id = session.get("manager_data", {}).get("manager_id")

    try:
        # Get manager info (sender email)
        manager = db.session.query(manager_info).filter_by(managerinfo_uid=manager_id).first()

        if not manager or not manager.manager_email:
            return jsonify({"success": False, "message": "Manager email not found"})

        # HR email list
        hr_emails = ["hr.admin@pfepl.com", "nisha.manikpurkar@pfepl.com"]

        # Compose email
        msg = Message(
            subject=f"Remark for {employee_name} - {month}", 
            recipients=hr_emails,               # Both HR emails in "To"
            cc=[manager.manager_email],         # Manager in "CC"
            body=f"""
Respected Ma'am,

I hope this message finds you well.

Please do the needful as per the following remark for the employee listed below for the month of {month}:

Employee Name : {employee_name}  
Month         : {month}  
Remark        : {remark}

Submitted by:  
Manager Name  : {manager.name}  
Manager Email : {manager.manager_email}

This is an automated message generated by the Invoice Management System.

Warm regards,  
Invoice Management System  
PFEPL
"""
        )

        mail.send(msg)

        # Optionally save to database
        if 'contractuserinfo_uid' in data:
            save_remark_to_database(data['contractuserinfo_uid'], data['month'], remark)

        return jsonify({"success": True, "message": "Email sent successfully"})

    except Exception as e:
        print(f"Error sending email: {str(e)}")
        return jsonify({"success": False, "message": f"Error sending email: {str(e)}"})


def save_remark_to_database(employee_uid, month, remark):
    try:
        # Update the remark in the invoice table
        db.session.execute(
            text("""
                UPDATE invoice_no_table 
                SET remarks = :remark
                WHERE contract_employee_uid = :employee_uid 
                AND month_date = :month
            """),
            {"remark": remark, "employee_uid": employee_uid, "month": month}
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error saving remark: {str(e)}")


if __name__ == "__main__":
    # For production, use gunicorn or waitress with proper timeout settings:
    # gunicorn --timeout 600 --workers 4 --bind 0.0.0.0:7071 app:app
    # OR for Windows: waitress-serve --port=7071 --channel-timeout=600 app:app
    
    import logging
    logging.basicConfig(level=logging.INFO)
    
    print("=" * 70)
    print("WARNING: Flask development server has limited timeout handling.")
    print("For production or heavy bulk uploads, use:")
    print("  - gunicorn (Linux/Mac): gunicorn --timeout 600 --workers 4 --bind 0.0.0.0:7071 app:app")
    print("  - waitress (Windows): waitress-serve --port=7071 --channel-timeout=600 app:app")
    print("=" * 70)
    
    app.run(debug=True, host="0.0.0.0", port=7071, threaded=True)

